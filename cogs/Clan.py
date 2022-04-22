from http.client import FORBIDDEN
from unittest import expectedFailure
import discord, openpyxl, time, discord_components, main
from discord.ext import commands
from discord_components import DiscordComponents, Select, SelectOption, Button, ButtonStyle, ComponentsBot
from main import bot
from openpyxl.utils import get_column_letter
import datetime

class commands(commands.Cog):
    def __init__(self, client):
        self.client = client

    @commands.command(name="clan_convidar")
    async def clan_convidar(self, ctx: commands.Context, member: discord.Member):

        member = member

        id = ctx.author.id
        member_id = "`" + str(id) + "`"

        file = "banco de dados.xlsx"
        wb = openpyxl.load_workbook(filename=file)
        ws = wb.worksheets[0]
        wsr = wb.active
        n = 0

        for row in wsr.iter_rows(wsr.min_row, wsr.max_row):
            for cell in row:
                if cell.value == member_id:
                    clan = ws.cell(row=cell.row, column=9).value
                    n += 1
                else:
                    n += 0

        if clan == "Não participante":
            embedVar = discord.Embed(title="⚠️Error 401", description="Você não participa de um clan", colour=discord.Colour.random())
            embedVar.set_author(name=ctx.author.name, icon_url=ctx.author.avatar_url)
            embedVar.set_footer(text="➝ Se precisar de ajuda, entre em contato com o DEV.")
            await ctx.reply(embed = embedVar)
            return

        if not "Líder" in clan:
            embedVar = discord.Embed(title="⚠️Error 403", description="Você não é líder do seu clan", colour=discord.Colour.random())
            embedVar.set_author(name=ctx.author.name, icon_url=ctx.author.avatar_url)
            embedVar.set_footer(text="➝ Se precisar de ajuda, entre em contato com o DEV.")
            await ctx.reply(embed=embedVar)
            return

        if n == 0:
            embedVar = discord.Embed(title="⚠️Error 404", description="Não foi encontrado o seu registro em nosso servidor!", colour=discord.Colour.random())
            embedVar.set_author(name=ctx.author.name, icon_url=ctx.author.avatar_url)
            embedVar.set_footer(text="Por favor, entre em contato com o DEV.")
            await ctx.send(embed = embedVar)
            return

        clan_ = clan[:-8]

        member_id = "`" + str(member.id) + "`"

        n = 0

        for row in wsr.iter_rows(wsr.min_row, wsr.max_row):
            for cell in row:
                if cell.value == member_id:
                    clan = ws.cell(row=cell.row, column=9).value
                    n += 1
                else:
                    n += 0

        if clan != "Não participante":
            embedVar = discord.Embed(title="⚠️Error 403", description=f"{member.name} já participa do clan {clan}")
            embedVar.set_author(name=ctx.author.name, icon_url=ctx.author.avatar_url)
            embedVar.set_footer(text="➝ Se precisar de ajuda, entre em contato com o DEV.")
            await ctx.send(embed = embedVar)
            return

        if n == 0:
            embedVar = discord.Embed(title="⚠️Error 404", description="Não foi encontrado o seu registro em nosso servidor!", colour=discord.Colour.random())
            embedVar.set_author(name=ctx.author.name, icon_url=ctx.author.avatar_url)
            embedVar.set_footer(text="Por favor, entre em contato com o DEV.")
            await ctx.send(embed = embedVar)
            return

        embedVar = discord.Embed(title="✅ Convidado com sucesso!", description=f"{member.name} foi convidado com sucesso para o seu clan.", colour = discord.Colour.random())
        embedVar.add_field(name="Observações:", value="➝ O convite pode expirar depois de um tempo.\n➝ O usuário em questão deve permitir receber mensagens de bot no privado do Discord.")
        embedVar.set_author(name=ctx.author.name, icon_url=ctx.author.avatar_url)
        embedVar.set_footer(text="➝ Se precisar de ajuda, entre em contato com o DEV.")
        response = await ctx.send(embed = embedVar)

        await ctx.invoke(self.client.get_command('clan_convidar_function'), member=member, clan=clan_)
        return

    @commands.command(name="clan_convidar_function")
    async def clan_convidar_function(self, ctx:commands.Context, member: discord.Member, *, clan):

        member_id = "`" + str(member.id) + "`"
        clan_ = clan

        try:
            embedVar = discord.Embed(title=f"⚜️ Convite para entrar no clan {clan_}", description=f"{ctx.author.name} te convidou para entrar no clan {clan_}. Para aceitar, clique no botão abaixo", colour=discord.Colour.random())
            embedVar.add_field(name="Observações:", value="・Você pode sair a hora que você quiser do clan digitando `!clan_sair` no servidor do State of War\n・Caso você não queira entrar no clan, apenas ignore.\n・Caso receba a mensagem dizendo que a interação falhou, significa que o convite já expirou")
            embedVar.set_author(name=ctx.author.name, icon_url=ctx.author.avatar_url)
            embedVar.set_footer(text="➝ Se precisar de ajuda, entre em contato com o DEV.")
            ask = await member.send(embed = embedVar, components=[Button(label="Entrar", style=ButtonStyle.green, custom_id="entrar")])

            await bot.wait_for('button_click', check=lambda i: i.custom_id == "entrar" and i.author == member)

        except Exception:
            embedVar = discord.Embed(title=f"<a:no_mark:963300068578963516> | Erro 403: Forbidden", description=f"Não foi possível convidar {member.name}, pois ele não permite receber mensagens de bots em seu privado.", colour=discord.Colour.red())
            embedVar.set_author(name=ctx.author.name, icon_url=ctx.author.avatar_url)
            embedVar.set_footer(text="➝ Se precisar de ajuda, entre em contato com o DEV.")
            await ctx.send(embed=embedVar)
            return

        while str(datetime.datetime.now().hour) == 0:
            continue

        file = "banco de dados.xlsx"
        wb = openpyxl.load_workbook(filename=file)
        ws = wb.worksheets[0]
        wsr = wb.active
        n = 0

        for row in wsr.iter_rows(wsr.min_row, wsr.max_row):
            for cell in row:
                if cell.value == member_id:
                    ws[f'{get_column_letter(cell.column + 8)}{(cell.row)}'] = f"{clan_}"
                    wb.save(file)

        for role in ctx.guild.roles:
            if role.name == f"{clan_}":
                await member.add_roles(role)
                await ask.delete()

        embedVar = discord.Embed(title=f"✔️ Parabéns! Você acaba de entrar no clan {clan_}!")
        embedVar.add_field(name="Observações:", value="・Você pode sair a hora que você quiser do clan digitando `!clan_sair` no servidor do State of War")
        embedVar.set_author(name=ctx.author.name, icon_url=ctx.author.avatar_url)
        embedVar.set_footer(text="➝ Se precisar de ajuda, entre em contato com o DEV.")
        await member.send(embed=embedVar)
        return

    @commands.command(name="clan_excluir")
    async def clan_excluir(self, ctx):

        if str(datetime.datetime.now().hour) == '8':
            warning = await ctx.reply("Estamos atualizando o nosso banco de dados. Por favor, tente novamente mais tarde\n_Esse processo dura das 00:00 até às 01:00_\nCaso você ache que isso seja um erro, entre em contato com o DEV.")
            time.sleep(10)
            await warning.delete()
            await ctx.message.delete()
            return

        id = ctx.author.id
        member_id = "`" + str(id) + "`"

        file = "banco de dados.xlsx"
        wb = openpyxl.load_workbook(filename=file)
        ws = wb.worksheets[0]
        wsr = wb.active
        n = 0

        for row in wsr.iter_rows(wsr.min_row, wsr.max_row):
            for cell in row:
                if cell.value == member_id:
                    clan = ws.cell(row=cell.row, column=9).value
                    n += 1
                else:
                    n += 0

        if clan == "Não participante":
            embedVar = discord.Embed(title="⚠️Error 401", description="Você não participa de um clan", colour=discord.Colour.random())
            embedVar.set_author(name=ctx.author.name, icon_url=ctx.author.avatar_url)
            embedVar.set_footer(text="➝ Se precisar de ajuda, entre em contato com o DEV.")
            await ctx.reply(embed = embedVar)
            return

        if not "Líder" in clan:
            embedVar = discord.Embed(title="⚠️Error 403", description="Você não é líder do seu clan", colour=discord.Colour.random())
            embedVar.set_author(name=ctx.author.name, icon_url=ctx.author.avatar_url)
            embedVar.set_footer(text="➝ Se precisar de ajuda, entre em contato com o DEV.")
            await ctx.reply(embed=embedVar)
            return

        clan = clan[:-8]
        await ctx.invoke(self.client.get_command('clan_excluir_function'), clan=clan)
        return

    @commands.command(name="clan_excluir_function")
    async def clan_excluir_function(self, ctx, clan):
        embedVar = discord.Embed(title="⚠️Atenção", description="Esse é um processo sem volta. Deseja mesmo excluir o seu clan? Todos os seus membros serão expulsos do seu clan.\nClique em 'Confirmar' para deletar o clan. Caso queira cancelar, apenas ignore.", colour=discord.Colour.random())
        embedVar.add_field(name="Observações:", value="・Apenas o autor do comando pode confirmar essa ação.\n・Essa ação é IRREVERSÍVEL, mesmo com o suporte da STAFF.\n・Você ainda poderá criar outro clan.")
        embedVar.set_author(name=ctx.author.name, icon_url=ctx.author.avatar_url)
        embedVar.set_footer(text="➝ Se precisar de ajuda, entre em contato com o DEV.")
        warning = await ctx.message.reply(embed=embedVar, components=[Button(label="Confirmar", style=ButtonStyle.green, custom_id="confirmado")])

        await bot.wait_for('button_click', check=lambda i: i.custom_id == "confirmado" and i.user == ctx.author)

        await warning.delete()
        embedVar = discord.Embed(title="Deletado com sucesso", description=f"O clan {clan} foi deletado com sucesso e todos seus membros foram desanexados.")
        embedVar.set_author(name=ctx.author.name, icon_url=ctx.author.avatar_url)
        embedVar.set_footer(text="➝ Se precisar de ajuda, entre em contato com o DEV.")
        await ctx.message.reply(embed=embedVar)

        for role in ctx.guild.roles:
            if role.name == f"{clan}":
                await role.delete()

        file = "banco de dados.xlsx"
        wb = openpyxl.load_workbook(filename=file)
        ws = wb.worksheets[0]
        wsr = wb.active
        n = 0

        for row in wsr.iter_cols(min_col=9, max_col=9, min_row=1):
                for cell in row:
                    if cell.value == f'{clan}' and cell.value != None:
                        ws[f'{get_column_letter(cell.column)}{(cell.row)}'] = "Não participante" 
                        wb.save(file)
                    elif cell.value == f'{clan}' + " (Líder)" and cell.value != None:
                        ws[f'{get_column_letter(cell.column)}{(cell.row)}'] = "Não participante" 
                        wb.save(file)
        return

def setup(client):
    client.add_cog(commands(client))

