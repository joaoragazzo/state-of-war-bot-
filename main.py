from ast import Pass
import asyncio
import functools
import itertools
import math
import random
import youtube_dl
from async_timeout import timeout
import nacl
import ffmpeg

import datetime 
import discord 
from discord import channel, colour, reaction 
from discord.ext import commands, tasks 
from discord.ext.commands.core import check, command 
from discord import File 
import discord_components 
from discord_components import DiscordComponents, Select, SelectOption, Button, ButtonStyle, ComponentsBot 
import openpyxl 
from openpyxl import load_workbook 
from openpyxl.utils import get_column_letter 
import time 
import requests
import datetime

intents = discord.Intents.all()
bot = commands.Bot(command_prefix=['!'], description="Bot", intents=intents)
DiscordComponents(bot)
client = discord.Client(intents=intents)

id_requests_chernarus = list()
id_channel_chernarus = list()

id_requests_namalsk = list()
id_channel_namalsk = list()

ids_suspeitos  = list()

youtube_dl.utils.bug_reports_message = lambda: ''

class VoiceError(Exception):
    pass

class YTDLError(Exception):
    pass

class YTDLSource(discord.PCMVolumeTransformer):
    YTDL_OPTIONS = {
        'format': 'bestaudio/best',
        'extractaudio': True,
        'audioformat': 'mp3',
        'outtmpl': '%(extractor)s-%(id)s-%(title)s.%(ext)s',
        'restrictfilenames': True,
        'noplaylist': True,
        'nocheckcertificate': True,
        'ignoreerrors': False,
        'logtostderr': False,
        'quiet': True,
        'no_warnings': True,
        'default_search': 'auto',
        'source_address': '0.0.0.0',
    }

    FFMPEG_OPTIONS = {
        'before_options': '-reconnect 1 -reconnect_streamed 1 -reconnect_delay_max 5',
        'options': '-vn',
    }

    ytdl = youtube_dl.YoutubeDL(YTDL_OPTIONS)

    def __init__(self, ctx: commands.Context, source: discord.FFmpegPCMAudio, *, data: dict, volume: float = 0.5):
        super().__init__(source, volume)

        self.requester = ctx.author
        self.channel = ctx.channel
        self.data = data

        self.uploader = data.get('uploader')
        self.uploader_url = data.get('uploader_url')
        date = data.get('upload_date')
        self.upload_date = date[6:8] + '.' + date[4:6] + '.' + date[0:4]
        self.title = data.get('title')
        self.thumbnail = data.get('thumbnail')
        self.description = data.get('description')
        self.duration = self.parse_duration(int(data.get('duration')))
        self.tags = data.get('tags')
        self.url = data.get('webpage_url')
        self.views = data.get('view_count')
        self.likes = data.get('like_count')
        self.dislikes = data.get('dislike_count')
        self.stream_url = data.get('url')

    def __str__(self):
        return '**{0.title}** por **{0.uploader}**'.format(self)

    @classmethod
    async def create_source(cls, ctx: commands.Context, search: str, *, loop: asyncio.BaseEventLoop = None):
        loop = loop or asyncio.get_event_loop()

        partial = functools.partial(cls.ytdl.extract_info, search, download=False, process=False)
        data = await loop.run_in_executor(None, partial)

        if data is None:
            raise YTDLError('Couldn\'t find anything that matches `{}`'.format(search))

        if 'entries' not in data:
            process_info = data
        else:
            process_info = None
            for entry in data['entries']:
                if entry:
                    process_info = entry
                    break

            if process_info is None:
                raise YTDLError('Couldn\'t find anything that matches `{}`'.format(search))

        webpage_url = process_info['webpage_url']
        partial = functools.partial(cls.ytdl.extract_info, webpage_url, download=False)
        processed_info = await loop.run_in_executor(None, partial)

        if processed_info is None:
            raise YTDLError('Couldn\'t fetch `{}`'.format(webpage_url))

        if 'entries' not in processed_info:
            info = processed_info
        else:
            info = None
            while info is None:
                try:
                    info = processed_info['entries'].pop(0)
                except IndexError:
                    raise YTDLError('Couldn\'t retrieve any matches for `{}`'.format(webpage_url))

        return cls(ctx, discord.FFmpegPCMAudio(info['url'], **cls.FFMPEG_OPTIONS), data=info)

    @staticmethod
    def parse_duration(duration: int):
        minutes, seconds = divmod(duration, 60)
        hours, minutes = divmod(minutes, 60)
        days, hours = divmod(hours, 24)

        duration = []
        if days > 0:
            duration.append('{} days'.format(days))
        if hours > 0:
            duration.append('{} hours'.format(hours))
        if minutes > 0:
            duration.append('{} minutes'.format(minutes))
        if seconds > 0:
            duration.append('{} seconds'.format(seconds))

        return ', '.join(duration)

class Song:
    __slots__ = ('source', 'requester')

    def __init__(self, source: YTDLSource):
        self.source = source
        self.requester = source.requester

    def create_embed(self):
        embed = (discord.Embed(title='Agora estou tocando',
                               description='```css\n{0.source.title}\n```'.format(self),
                               color=discord.Color.red())
                 .add_field(name='A duração da música é de', value=self.source.duration)
                 .add_field(name='Essa música foi pedida por', value=self.requester.mention)
                 .add_field(name='O autor dessa música é', value='[{0.source.uploader}]({0.source.uploader_url})'.format(self))
                 .add_field(name='URL', value='[Click]({0.source.url})'.format(self))
                 .set_thumbnail(url=self.source.thumbnail))

        return embed

class SongQueue(asyncio.Queue):
    def __getitem__(self, item):
        if isinstance(item, slice):
            return list(itertools.islice(self._queue, item.start, item.stop, item.step))
        else:
            return self._queue[item]

    def __iter__(self):
        return self._queue.__iter__()

    def __len__(self):
        return self.qsize()

    def clear(self):
        self._queue.clear()

    def shuffle(self):
        random.shuffle(self._queue)

    def remove(self, index: int):
        del self._queue[index]

class VoiceState:
    def __init__(self, bot: commands.Bot, ctx: commands.Context):
        self.bot = bot
        self._ctx = ctx

        self.current = None
        self.voice = None
        self.next = asyncio.Event()
        self.songs = SongQueue()

        self._loop = False
        self._volume = 0.5
        self.skip_votes = set()

        self.audio_player = bot.loop.create_task(self.audio_player_task())

    def __del__(self):
        self.audio_player.cancel()

    @property
    def loop(self):
        return self._loop

    @loop.setter
    def loop(self, value: bool):
        self._loop = value

    @property
    def volume(self):
        return self._volume

    @volume.setter
    def volume(self, value: float):
        self._volume = value

    @property
    def is_playing(self):
        return self.voice and self.current

    async def audio_player_task(self):
        while True:
            self.next.clear()

            if not self.loop:
                # Try to get the next song within 15 minutes.
                # If no song will be added to the queue in time,
                # the player will disconnect due to performance
                # reasons.
                try:
                    async with timeout(900):  # 15 minutes
                        self.current = await self.songs.get()
                except asyncio.TimeoutError:
                    self.bot.loop.create_task(self.stop())
                    return

            self.current.source.volume = self._volume
            self.voice.play(self.current.source, after=self.play_next_song)
            await self.current.source.channel.send(embed=self.current.create_embed())

            await self.next.wait()

    def play_next_song(self, error=None):
        if error:
            raise VoiceError(str(error))

        self.next.set()

    def skip(self):
        self.skip_votes.clear()

        if self.is_playing:
            self.voice.stop()

    async def stop(self):
        self.songs.clear()

        if self.voice:
            await self.voice.disconnect()
            self.voice = None

class Music(commands.Cog):

    def __init__(self, bot: commands.Bot):
        self.bot = bot
        self.voice_states = {}

    def get_voice_state(self, ctx: commands.Context):
        state = self.voice_states.get(ctx.guild.id)
        if not state:
            state = VoiceState(self.bot, ctx)
            self.voice_states[ctx.guild.id] = state

        return state

    def cog_unload(self):
        for state in self.voice_states.values():
            self.bot.loop.create_task(state.stop())

    def cog_check(self, ctx: commands.Context):
        if not ctx.guild:
            raise commands.NoPrivateMessage('Esses comandos não funcionam nas DMs, por favor, utilize os comandos no nosso servidor.')

        return True

    async def cog_before_invoke(self, ctx: commands.Context):
        ctx.voice_state = self.get_voice_state(ctx)

    async def cog_command_error(self, ctx: commands.Context, error: commands.CommandError):
        await ctx.send('Um erro aconteceu, por favor, encaminhar esse código para o DEV: {}'.format(str(error)))

    @commands.command(name='entrar', invoke_without_subcommand=True)
    async def _join(self, ctx: commands.Context):

        destination = ctx.author.voice.channel
        if ctx.voice_state.voice:
            await ctx.voice_state.voice.move_to(destination)
            return

        ctx.voice_state.voice = await destination.connect()

    @commands.command(name='conectar')
    @commands.has_permissions(manage_guild=True)
    async def _summon(self, ctx: commands.Context, *, channel: discord.VoiceChannel = None):

        if not channel and not ctx.author.voice:
            raise VoiceError('Você não especificou nenhum canal de voz para eu conectar.')

        destination = channel or ctx.author.voice.channel
        if ctx.voice_state.voice:
            await ctx.voice_state.voice.move_to(destination)
            return

        ctx.voice_state.voice = await destination.connect()

    @commands.command(name='sair', aliases=['desconectar'])
    async def _leave(self, ctx: commands.Context):

        if not ctx.voice_state.voice:
            return await ctx.send('Eu não estou cantando em nenhum canal de voz.')

        await ctx.voice_state.stop()
        del self.voice_states[ctx.guild.id]

    @commands.command(name='agora', aliases=['current', 'tocando'])
    async def _now(self, ctx: commands.Context):
        await ctx.send(embed=ctx.voice_state.current.create_embed())

    @commands.command(name='pausar', aliases =["pause"])
    @commands.has_permissions(manage_guild=True)
    async def _pause(self, ctx: commands.Context):

        if not ctx.voice_state.is_playing and ctx.voice_state.voice.is_playing():
            ctx.voice_state.voice.pause()
            await ctx.message.add_reaction('✅')

    @commands.command(name='continuar', aliases=["resume"])
    @commands.has_permissions(manage_guild=True)
    async def _resume(self, ctx: commands.Context):

        if not ctx.voice_state.is_playing and ctx.voice_state.voice.is_paused():
            ctx.voice_state.voice.resume()
            await ctx.message.add_reaction('✅')

    @commands.command(name='reiniciar', aliases=["limpar"])
    @commands.has_permissions(manage_guild=True)
    async def _stop(self, ctx: commands.Context):

        ctx.voice_state.songs.clear()

        if not ctx.voice_state.is_playing:
            ctx.voice_state.voice.stop()
            await ctx.message.add_reaction('✅')

    @commands.command(name='s', aliases=["skip"])
    async def _skip(self, ctx: commands.Context):

        if not ctx.voice_state.is_playing:
            return await ctx.send('Eu não estou cantando nenhuma música atualmente.')

        voter = ctx.message.author
        if voter == ctx.voice_state.current.requester:
            await ctx.message.add_reaction('✅')
            ctx.voice_state.skip()

        if not voter == ctx.voice_state.current.requester:
            await ctx.message.add_reaction('✅')
            ctx.voice_state.skip()

        else:
            await ctx.send('Pulando para a próxima música!')

    @commands.command(name='fila')
    async def _queue(self, ctx: commands.Context, *, page: int = 1):

        if len(ctx.voice_state.songs) == 0:
            return await ctx.send('Não tem nenhuma música na fila, ainda...')

        items_per_page = 10
        pages = math.ceil(len(ctx.voice_state.songs) / items_per_page)

        start = (page - 1) * items_per_page
        end = start + items_per_page

        queue = ''
        for i, song in enumerate(ctx.voice_state.songs[start:end], start=start):
            queue += '`{0}.` [**{1.source.title}**]({1.source.url})\n'.format(i + 1, song)

        embed = (discord.Embed(description='**{} músicas na fila:**\n\n{}'.format(len(ctx.voice_state.songs), queue))
                 .set_footer(text='Vendo a página {}/{}'.format(page, pages)))
        await ctx.send(embed=embed)

    @commands.command(name='embaralhar')
    async def _shuffle(self, ctx: commands.Context):

        if len(ctx.voice_state.songs) == 0:
            return await ctx.send('```Impossível realizar essa ação: não existem músicas na fila.```')

        ctx.voice_state.songs.shuffle()
        await ctx.message.add_reaction('✅')

    @commands.command(name='remover')
    async def _remove(self, ctx: commands.Context, index: int):

        if len(ctx.voice_state.songs) == 0:
            return await ctx.send('Não quero cantar nenhuma musica ainda')

        ctx.voice_state.songs.remove(index - 1)
        await ctx.message.add_reaction('✅')

    @commands.command(name='loop')
    async def _loop(self, ctx: commands.Context):

        if not ctx.voice_state.is_playing:
            return await ctx.send('Nada está sendo tocado no momento.')

        # Inverse boolean value to loop and unloop.
        ctx.voice_state.loop = not ctx.voice_state.loop
        await ctx.message.add_reaction('✅')

    @commands.command(name='p', aliases=["play"])
    async def _play(self, ctx: commands.Context, *, search: str):

        if not ctx.voice_state.voice:
            await ctx.invoke(self._join)

        async with ctx.typing():
            try:
                source = await YTDLSource.create_source(ctx, search, loop=self.bot.loop)
            except YTDLError as e:
                await ctx.send('Ocorreu um erro enquanto eu processava o seu pedido: {}'.format(str(e)))
            else:
                song = Song(source)

                await ctx.voice_state.songs.put(song)
                await ctx.send('{} foi adicionado na fila'.format(str(source)))

    @_join.before_invoke
    @_play.before_invoke
    async def ensure_voice_state(self, ctx: commands.Context):
        if not ctx.author.voice or not ctx.author.voice.channel:
            raise commands.CommandError('Você não está conectado em nenhum chat de voz!')

        if ctx.voice_client:
            if ctx.voice_client.channel != ctx.author.voice.channel:
                raise commands.CommandError('Eu já estou cantando em outro chat de voz!')    

@bot.command(name="suspeito_adicionar")
async def suspeito_adicionar(ctx, *,conteudo):

    check_1 = discord.utils.get(ctx.guild.roles, id=947242380040478761)

    if check_1 in ctx.author.roles:
        pass
    else:
        warning = await ctx.reply("Você não tem permissão para usar isso!")
        time.sleep(8)
        await warning.delete()
        await ctx.message.delete()
        return

    conteudo = conteudo.split("/")
    nick = conteudo[0]
    id = conteudo[1]

    if not '76' in id:
        embedVar = discord.Embed(title="⚠️ SteamID não válido!", description="Por favor, certifique-se que você utilizou a forma `!suspeito_adicionar [NICK]/[STEAMID]/[MOTIVO]`")
        embedVar.set_footer(text="Caso tenha alguma duvida, entre em contato com  DEV.")
        await ctx.send(embed=embedVar)
        return


    id = "`" + str(id) + "`"
    motivo = conteudo[2]

    valid_reaction=['✅','❌']

    file = "suspeitos.xlsx"
    wb = openpyxl.load_workbook(filename=file)
    ws = wb.worksheets[0]
    wsr = wb.active

    for row in wsr.iter_rows(wsr.min_row, wsr.max_row):
        for cell in row:
            if cell.value == id:
                suspeito_already = discord.Embed(title="⚠️ Esse SteamID já encontrado!", description="O SteamID já foi encontrado no nosso banco de dados. Deseja ver mais informações", colour=discord.Colour.red())
                suspeito_already.set_footer(text="Caso tenha alguma dúvida, contate o DEV.")
                warning = await ctx.send(embed=suspeito_already)

                await warning.add_reaction('✅')
                await warning.add_reaction('❌')

                def check(reaction, user):
                    return user == ctx.author and str(reaction.emoji) in valid_reaction
                reaction, user = await bot.wait_for('reaction_add', timeout=60.0, check=check)

                if str(reaction.emoji) == '✅':
                    await warning.delete()
                    id = ws.cell(row=cell.row, column = 1).value
                    nick = ws.cell(row=cell.row, column = 2).value
                    motivo = ws.cell(row=cell.row, column = 3).value

                    embedVar = discord.Embed(title=f"⚠️ Suspeito encontrado com sucesso!", description=f"**SteamID:** {id}\n**Nick:** {nick}\n**Motivo:** {motivo}", colour = discord.Colour.red())
                    embedVar.set_footer(text="Caso tenha alguma dúvida, contate o DEV.")
                    await ctx.send(embed = embedVar)
                    return 
                
                elif str(reaction.emoji) == '❌':
                    await warning.delete()
                    return        
                
    ws[f'{get_column_letter(ws.min_column + 0)}{(ws.max_row + 1)}'] = id
    ws[f'{get_column_letter(ws.min_column + 1)}{(ws.max_row)}'] = nick
    ws[f'{get_column_letter(ws.min_column + 2)}{(ws.max_row)}'] = motivo

    embedVar = discord.Embed(title="⚠️ Adicionado com sucesso!", description=f"O jogador `{nick}`, portador do SteamID `{id}` foi adicionado na lista de suspeitos por `{motivo}`.",colour = discord.Colour.red())
    embedVar.set_footer(text="Caso tenha alguma dúvida, contate o DEV.")
    await ctx.send(embed=embedVar)

    wb.save(file)
    return

@bot.command(name="suspeito_editar")
async def lista(ctx, id):

    check_1 = discord.utils.get(ctx.guild.roles, id=947242380040478761)

    if check_1 in ctx.author.roles:
        pass
    else:
        warning = await ctx.reply("Você não tem permissão para usar isso!")
        time.sleep(8)
        await warning.delete()
        await ctx.message.delete()
        return

    id = "`" + str(id) + "`"

    valid_reaction=["1️⃣","2️⃣","3️⃣"]

    file = "suspeitos.xlsx"
    wb = openpyxl.load_workbook(filename=file)
    ws = wb.worksheets[0]
    wsr = wb.active

    for row in wsr.iter_rows(wsr.min_row, wsr.max_row):
        for cell in row:
            if cell.value == id:
                embedVar = discord.Embed(title=f"Atualização de suspeita do ID {id}", description="O que você deseja editar?\n1️⃣ ID\n2️⃣Nick\n3️⃣Motivo da suspeita", colour = discord.Colour.red())
                embedVar.set_footer(text="Caso tenha alguma dúvida, entre em contato com o DEV.")
                mensagem = await ctx.send(embed=embedVar)

                def check(reaction, user):
                    return user == ctx.author and str(reaction.emoji) in valid_reaction
                reaction, user = await bot.wait_for('reaction_add', timeout = 60, check=check)

                def check_msg(msg):
                        return msg.author == ctx.author

                if str(reaction.emoji) == '1️⃣':
                    await mensagem.delete()
                    embedVar = discord.Embed(title=f"Atualização do ID do suspeito", description="Por favor, digite o novo ID que você queira colocar para o suspeito", colour = discord.Colour.red())
                    embedVar.set_footer(text="Caso tenha alguma dúvida, entre em contato com o DEV.")
                    mensagem = await ctx.send(embed=embedVar)

                    new_ID = await bot.wait_for('message', timeout=60, check=check_msg)
                    new_ID = "`" + str(new_ID.content) + "`"
                    ws[cell.coordinate] = new_ID
                    wb.save(file)

                    await mensagem.delete()
                    await new_ID.delete()

                    embedVar = discord.Embed(title=f"Alterado com sucesso!", description=f"O ID do suspeito foi alterado com sucesso de {id} para {new_ID}!", colour = discord.Colour.red())
                    embedVar.set_footer(text="Caso tenha alguma dúvida, entre em contato com o DEV.")
                    await ctx.send(embed=embedVar)

                elif str(reaction.emoji) == '2️⃣':
                    await mensagem.delete()
                    embedVar = discord.Embed(title=f"Atualização do NICK do suspeito", description="Por favor, digite o novo nick que você queira colocar para o suspeito", colour = discord.Colour.red())
                    embedVar.set_footer(text="Caso tenha alguma dúvida, entre em contato com o DEV.")
                    mensagem = await ctx.send(embed=embedVar)

                    nick = await bot.wait_for('message', timeout=60, check=check_msg)
                    nick = nick.content
                    ws[cell.coordinate] = nick
                    wb.save(file)

                    await mensagem.delete()
                    await nick.delete()

                    embedVar = discord.Embed(title=f"Alterado com sucesso!", description=f"O ID do suspeito foi alterado com sucesso para {nick}!", colour = discord.Colour.red())
                    embedVar.set_footer(text="Caso tenha alguma dúvida, entre em contato com o DEV.")
                    await ctx.send(embed=embedVar)

                elif str(reaction.emoji) == '3️⃣':
                    await mensagem.delete()
                    embedVar = discord.Embed(title=f"Atualização do motivo de suspeita", description="Por favor, digite o novo motivo que você queira colocar para o suspeito", colour = discord.Colour.red())
                    embedVar.set_footer(text="Caso tenha alguma dúvida, entre em contato com o DEV.")
                    mensagem = await ctx.send(embed=embedVar)

                    motivo = await bot.wait_for('message', timeout=60, check=check_msg)
                    motivo = motivo.content
                    ws[cell.coordinate] = motivo
                    wb.save(file)

                    await mensagem.delete()
                    await motivo.delete()

                    embedVar = discord.Embed(title=f"Alterado com sucesso!", description=f"O ID do suspeito foi alterado com sucesso para {motivo}!", colour = discord.Colour.red())
                    embedVar.set_footer(text="Caso tenha alguma dúvida, entre em contato com o DEV.")
                    await ctx.send(embed=embedVar)

@bot.command(name="suspeito_lista")
async def suspeito_lista(ctx):

    check_1 = discord.utils.get(ctx.guild.roles, id=947242380040478761)

    if check_1 in ctx.author.roles:
        pass
    else:
        warning = await ctx.reply("Você não tem permissão para usar isso!")
        time.sleep(8)
        await warning.delete()
        await ctx.message.delete()
        return
    
    file = "suspeitos.xlsx"
    wb = openpyxl.load_workbook(filename=file)
    ws = wb.worksheets[0]
    lista = ""
    
    n = 0
    for row in range(2, ws.max_row+1):
        for column in "A":
            n = n + 1
            cell_name="{}{}".format(column, row)
            lista = lista + f"{n}. {ws[cell_name].value}\n"
    
    lista = discord.Embed(title="Lista de todos os SteamIDs suspeitos:", description=f"{lista}", color = discord.Colour.random())
    lista.set_footer(text="Para mais informações sobre o motivo de report do ID, utilize !suspeito_ficha")
    await ctx.reply(embed=lista)

@bot.command(name="suspeito_ficha")
async def suspeito_ficha(ctx, id):

    check_1 = discord.utils.get(ctx.guild.roles, id=947242380040478761)

    if check_1 in ctx.author.roles:
        pass
    else:
        warning = await ctx.reply("Você não tem permissão para usar isso!")
        time.sleep(8)
        await warning.delete()
        await ctx.message.delete()
        return

    id = "`" + str(id) + "`"

    file = "suspeitos.xlsx"
    wb = openpyxl.load_workbook(filename=file)
    ws = wb.worksheets[0]
    wsr = wb.active

    n = 0 

    for row in wsr.iter_rows(wsr.min_row, wsr.max_row):
        for cell in row:
            if cell.value == id:
                n += 1
                nick = ws.cell(row=cell.row, column = 2).value
                motivo = ws.cell(row=cell.row, column = 3).value
    
    if n == 0:
        embedVar = discord.Embed(title="⚠️ Não encontrado!", description="O SteamID em questão não foi encontrado no nosso banco de dados.", colour = discord.Colour.red())
        embedVar.set_footer(text="Caso tenha alguma dúvida, entre em contato com o DEV.")
        await ctx.send(embed = embedVar)
        return

    embedVar = discord.Embed(title="⚠️ Suspeito encontrado com sucesso!", description=f"**SteamID:** {id}\n**Nick:** {nick}\n**Motivo de suspeita:** {motivo}", colour=discord.Colour.red())
    embedVar.set_footer(text="Caso tenha alguma dúvida, entre em contato com o DEV.")
    await ctx.send(embed = embedVar)
    return   

@bot.command(name="suspeito_deletar")
async def suspeito_deletar(ctx, id):
    check_1 = discord.utils.get(ctx.guild.roles, id=947242380040478761)

    if check_1 in ctx.author.roles:
        pass
    else:
        warning = await ctx.reply("Você não tem permissão para usar isso!")
        time.sleep(8)
        await warning.delete()
        await ctx.message.delete()
        return

    id = "`" + str(id) + "`"

    file = "suspeitos.xlsx"
    wb = openpyxl.load_workbook(filename=file)
    ws = wb.worksheets[0]
    wsr = wb.active

    valid_reaction = ['✅','❌']

    n = 0 

    for row in wsr.iter_rows(wsr.min_row, wsr.max_row):
        for cell in row:
            if cell.value == id:
                n += 1
                embedVar = discord.Embed(title=f"Você deseja mesmo deletar a suspeita do SteamID `{id}`?", colour= discord.Colour.red())
                embedVar.set_footer(text="Caso tenha alguma dúvida, entre em contato com DEV.")
                message = await ctx.send(embed = embedVar)
                await message.add_reaction('✅')
                await message.add_reaction('❌')

                def check(reaction, user):
                    return user == ctx.author and str(reaction.emoji) in valid_reaction
                reaction, user = await bot.wait_for('reaction_add', timeout=60.0, check=check)

                if str(reaction.emoji) == '✅':
                    await message.delete()
                    ws[cell.coordinate] = "DELETADO"
                    wb.save(file)
                    embedVar = discord.Embed(title="✅ Deletado com sucesso!", colour = discord.Colour.red())
                    embedVar.set_footer(text="Caso tenha alguma dúvida, entre em contato com o DEV.")
                    await ctx.send(embed=embedVar)
                    return
                
                elif str(reaction.emoji) == '❌':
                    await message.delete()
                    return
    
    if n == 0:
        embedVar = discord.Embed(title="⚠️ Esse SteamID não foi encontrado na nossa lista!", description="Certifique-se que você digitou o SteamID corretamente!", colour=discord.Colour.red())
        embedVar.set_footer(text="Caso tenha alguma dúvida, entre em contato com o DEV.")
        await ctx.send(embed = embedVar)
                
@bot.command(name="purge")
async def purge(ctx):
    check_1 = discord.utils.get(ctx.guild.roles, id=947242380040478761)

    if check_1 in ctx.author.roles:
        pass
    else:
        warning = await ctx.reply("Você não tem permissão para usar isso!")
        time.sleep(8)
        await warning.delete()
        await ctx.message.delete()
        return

    warning = await ctx.send('Apagando todas as mensagens desse canal...')
    time.sleep(3)
    await ctx.message.delete()
    await warning.delete()
    await ctx.channel.purge()

@bot.command(name="kdr")
async def kdr(ctx):
    discord_id = ctx.author.id
    discord_id = "`" + str(discord_id) + "`"

    embedVar = discord.Embed(title="⚜️ Selecione o mapa", description="Por favor, selecione o mapa que você deseja olhar as suas estatísticas!\n☢️ Chernarus\n❄️ Namalsk", colour=discord.Colour.random())
    embedVar.set_footer(text="Caso tenha alguma dúvida, entre em contato com o DEV.")
    message = await ctx.reply(embed=embedVar)
    await message.add_reaction('☢️')
    await message.add_reaction('❄️')

    valid_reaction = ['☢️','❄️']

    def check(reaction, user):
        return user == ctx.author and str(reaction.emoji) in valid_reaction
    reaction, user = await bot.wait_for('reaction_add', timeout=60.0, check=check)

    if str(reaction.emoji) == '☢️':
        await message.delete()
        file = "kds_chernarus.xlsx"
        wb = openpyxl.load_workbook(filename=file)
        ws = wb.worksheets[0]
        wsr = wb.active

        n = 0

        for row in wsr.iter_rows(wsr.min_row, wsr.max_row):
            for cell in row:
                if cell.value == discord_id:
                    n += 1
                    mortes = ws.cell(row=cell.row, column = 3).value
                    kills = ws.cell(row=cell.row, column = 4).value
                    maior_distancia = ws.cell(row=cell.row, column = 5).value

        if n == 0:
            embedVar = discord.Embed(title="⚠️ Error 404", description="O seu usuário não foi encontrado no nosso banco de dados. Certifique-se de que você foi devidamente cadastrado na aba <#960654509904904252>", colour=discord.Colour.red())
            embedVar.set_footer(text="Caso tenha alguma dúvida, entre em contato com o DEV.")
            await ctx.send(embed=embedVar)
            return

        if int(mortes) != 0:
            kdr = int(kills)/int(mortes)
        else:
            kdr = int(kills)/1

        embedVar = discord.Embed(title="☠️ Kill Death Ratio | State of War | ☢️ Chernarus", description=f"<:killicon:965991780824797224> Kills: `{kills}`\n<:death:965992257301938249> Mortes: `{mortes}`\n<:kdr:965996010654548039> Kill Death Ratio (KDR): `{kdr}`\n<:sniper:965991798063370251> Maior distância de uma kill: `{maior_distancia} metros`", colour=discord.Colour.random())
        embedVar.set_footer(text="Caso tenha alguma dúvida, entre em contato com o DEV.")
        await ctx.reply(embed=embedVar)

    elif str(reaction.emoji) == '❄️':
        await message.delete()
        file = "kds_namalsk.xlsx"
        wb = openpyxl.load_workbook(filename=file)
        ws = wb.worksheets[0]
        wsr = wb.active

        n = 0

        for row in wsr.iter_rows(wsr.min_row, wsr.max_row):
            for cell in row:
                if cell.value == discord_id:
                    n += 1
                    mortes = ws.cell(row=cell.row, column = 3).value
                    kills = ws.cell(row=cell.row, column = 4).value
                    maior_distancia = ws.cell(row=cell.row, column = 5).value

        if n == 0:
            embedVar = discord.Embed(title="⚠️ Error 404", description="O seu usuário não foi encontrado no nosso banco de dados. Certifique-se de que você foi devidamente cadastrado na aba <#960654509904904252>", colour=discord.Colour.red())
            embedVar.set_footer(text="Caso tenha alguma dúvida, entre em contato com o DEV.")
            await ctx.send(embed=embedVar)
            return

        if int(mortes) != 0:
            kdr = int(kills)/int(mortes)
        else:
            kdr = int(kills)/1

        embedVar = discord.Embed(title="☠️ Kill Death Ratio | State of War | ❄️ Namalsk", description=f"<:killicon:965991780824797224> Kills: `{kills}`\n<:death:965992257301938249> Mortes: `{mortes}`\n<:kdr:965996010654548039> Kill Death Ratio (KDR): `{kdr}`\n<:sniper:965991798063370251> Maior distância de uma kill: `{maior_distancia} metros`", colour=discord.Colour.random())
        embedVar.set_footer(text="Caso tenha alguma dúvida, entre em contato com o DEV.")
        await ctx.reply(embed=embedVar)

@bot.command(name="kdr_pv")
async def kdr(ctx):
    discord_id = ctx.author.id
    discord_id = "`" + str(discord_id) + "`"

    embedVar = discord.Embed(title="⚜️ Selecione o mapa", description="Por favor, selecione o mapa que você deseja olhar as suas estatísticas!\n☢️ Chernarus\n❄️ Namalsk", colour=discord.Colour.random())
    embedVar.set_footer(text="Caso tenha alguma dúvida, entre em contato com o DEV.")
    message = await ctx.reply(embed=embedVar)
    await message.add_reaction('☢️')
    await message.add_reaction('❄️')

    valid_reaction = ['☢️','❄️']

    def check(reaction, user):
        return user == ctx.author and str(reaction.emoji) in valid_reaction
    reaction, user = await bot.wait_for('reaction_add', timeout=60.0, check=check)

    if str(reaction.emoji) == '☢️':
        file = "kds_chernarus.xlsx"
        wb = openpyxl.load_workbook(filename=file)
        ws = wb.worksheets[0]
        wsr = wb.active

        n = 0

        for row in wsr.iter_rows(wsr.min_row, wsr.max_row):
            for cell in row:
                if cell.value == discord_id:
                    n += 1
                    mortes = ws.cell(row=cell.row, column = 3).value
                    kills = ws.cell(row=cell.row, column = 4).value
                    maior_distancia = ws.cell(row=cell.row, column = 5).value

        if n == 0:
            embedVar = discord.Embed(title="⚠️ Error 404", description="O seu usuário não foi encontrado no nosso banco de dados. Certifique-se de que você foi devidamente cadastrado na aba <#960654509904904252>", colour=discord.Colour.red())
            embedVar.set_footer(text="Caso tenha alguma dúvida, entre em contato com o DEV.")
            await ctx.send(embed=embedVar)
            return

        if int(mortes) != 0:
            kdr = int(kills)/int(mortes)
        else:
            kdr = int(kills)/1

        embedVar = discord.Embed(title="☠️ Kill Death Ratio | State of War", description=f"<:killicon:965991780824797224> Kills: `{kills}`\n<:death:965992257301938249> Mortes: `{mortes}`\n<:kdr:965996010654548039> Kill Death Ratio (KDR): `{kdr}`\n<:sniper:965991798063370251> Maior distância de uma kill: `{maior_distancia} metros`", colour=discord.Colour.random())
        embedVar.set_footer(text="Caso tenha alguma dúvida, entre em contato com o DEV.")
        await ctx.author.send(embed=embedVar)
    
    elif str(reaction.emoji) == '❄️':
        file = "kds_namalsk.xlsx"
        wb = openpyxl.load_workbook(filename=file)
        ws = wb.worksheets[0]
        wsr = wb.active

        n = 0

        for row in wsr.iter_rows(wsr.min_row, wsr.max_row):
            for cell in row:
                if cell.value == discord_id:
                    n += 1
                    mortes = ws.cell(row=cell.row, column = 3).value
                    kills = ws.cell(row=cell.row, column = 4).value
                    maior_distancia = ws.cell(row=cell.row, column = 5).value

        if n == 0:
            embedVar = discord.Embed(title="⚠️ Error 404", description="O seu usuário não foi encontrado no nosso banco de dados. Certifique-se de que você foi devidamente cadastrado na aba <#960654509904904252>", colour=discord.Colour.red())
            embedVar.set_footer(text="Caso tenha alguma dúvida, entre em contato com o DEV.")
            await ctx.send(embed=embedVar)
            return

        if int(mortes) != 0:
            kdr = int(kills)/int(mortes)
        else:
            kdr = int(kills)/1

        embedVar = discord.Embed(title="☠️ Kill Death Ratio | State of War", description=f"<:killicon:965991780824797224> Kills: `{kills}`\n<:death:965992257301938249> Mortes: `{mortes}`\n<:kdr:965996010654548039> Kill Death Ratio (KDR): `{kdr}`\n<:sniper:965991798063370251> Maior distância de uma kill: `{maior_distancia} metros`", colour=discord.Colour.random())
        embedVar.set_footer(text="Caso tenha alguma dúvida, entre em contato com o DEV.")
        await ctx.author.send(embed=embedVar)

@bot.command(name="top_kills")
async def top_kills(ctx):

    valid_reaction=["☢️","❄️"]

    embedVar = discord.Embed(title="⚜️ Selecione o mapa", description="Por favor, selecione o mapa que você deseja olhar as melhores kills!\n☢️ Chernarus\n❄️ Namalsk", colour=discord.Colour.random())
    embedVar.set_footer(text="Caso tenha alguma dúvida, entre em contato com o DEV.")
    message = await ctx.send(embed = embedVar)

    await message.add_reaction('☢️')
    await message.add_reaction('❄️')

    def check(reaction, user):
        return user == ctx.author and str(reaction.emoji) in valid_reaction
    reaction, user = await bot.wait_for('reaction_add', timeout=60.0, check=check)

    if str(reaction.emoji) == '☢️':
        file = "kds_chernarus.xlsx"
        wb = openpyxl.load_workbook(filename=file)
        ws = wb.worksheets[0]
        wsr = wb.active

        inicial_list = []
        medium_list = []
        final_list = []

        for row in wsr.iter_cols(min_col=5, max_col=5, min_row=1):
            for cell in row:
                if cell.value != None:
                    distancia = cell.value
                    inicial_list.append(distancia)

        for dis in inicial_list:
            dis = float(dis)
            medium_list.append(dis)

        medium_list.sort(reverse=True)
        for bests in medium_list[:10]:
            final_list.append(bests)

        top_kills_message = ""

        number = 1

        melhores_jogadores = []

        for n in final_list:
            for row in wsr.iter_cols(min_col=5, max_col=5, min_row=1):
                for cell in row:
                    if cell.value == n and cell.value != None:
                        discord_id_bests = ws.cell(row=cell.row, column = 1).value
                        discord_id_bests = discord_id_bests.replace("`","")
                        discord_id_bests = int(discord_id_bests)
                        melhores_jogadores.append(discord_id_bests)

        for n in final_list:
            i = final_list.index(n)
            member = await bot.fetch_user(melhores_jogadores[i])
            top_kills_message = top_kills_message + str(str(number) +"º. " + "`" + str(n) + "`" + " metros. Killer: " + str((member.display_name)) + "\n")
            number += 1

        embedVar = discord.Embed(title="<:sniper:965991798063370251> Melhores 10 tiros State of War", description=top_kills_message)
        embedVar.set_footer(text="Caso tenha alguma dúvida, entre em contato com o DEV.")
        await ctx.send(embed=embedVar)    

    elif str(reaction.emoji) == '❄️':
        file = "kds_namalsk.xlsx"
        wb = openpyxl.load_workbook(filename=file)
        ws = wb.worksheets[0]
        wsr = wb.active

        inicial_list = []
        medium_list = []
        final_list = []

        for row in wsr.iter_cols(min_col=5, max_col=5, min_row=1):
            for cell in row:
                if cell.value != None:
                    distancia = cell.value
                    inicial_list.append(distancia)

        for dis in inicial_list:
            dis = float(dis)
            medium_list.append(dis)

        medium_list.sort(reverse=True)

        for bests in medium_list[:10]:
            final_list.append(bests)

        top_kills_message = ""
        
        number = 1

        melhores_jogadores = []

        for n in final_list:
            for row in wsr.iter_cols(min_col=5, max_col=5, min_row=1):
                for cell in row:
                    if str(cell.value) == str(n) and cell.value != None:
                        discord_id_bests = ws.cell(row=cell.row, column = 1).value
                        discord_id_bests = discord_id_bests.replace("`","")
                        discord_id_bests = int(discord_id_bests)
                        melhores_jogadores.append(discord_id_bests)

        for n in final_list:
            i = final_list.index(n)
            member = await bot.fetch_user(melhores_jogadores[i])
            top_kills_message = top_kills_message + str(str(number) +"º. " + "`" + str(n) + "`" + " metros. Killer: " + str(member.display_name) + "\n")
            number += 1

        embedVar = discord.Embed(title="<:sniper:965991798063370251> Melhores 10 tiros State of War", description=top_kills_message)
        embedVar.set_footer(text="Caso tenha alguma dúvida, entre em contato com o DEV.")
        await ctx.send(embed=embedVar)         

@bot.command(name="regras_1")
async def regras_1(ctx):   
    check_1 = discord.utils.get(ctx.guild.roles, id=947242380040478761)

    if check_1 in ctx.author.roles:
        pass
    else:
        warning = await ctx.reply("Você não tem permissão para usar isso!")
        time.sleep(8)
        await warning.delete()
        await ctx.message.delete()
        return

    await ctx.message.delete()

    await ctx.send(file=discord.File("images/regras_chernarus.png"))
    embedVar = discord.Embed(title="**Regras do servidor**", description="É estritamente proibido, dentro das dependências do State of War:",colour=discord.Colour.red())
    embedVar.set_footer(text="As regras estão sujeitas a mudanças. Fique atento!")
    await ctx.send(embed=embedVar)
    await ctx.send("```python\n1・Uso de programas externos que possam oferecer vantagem dentro do jogo (hacks/sem mato);\n2・Uso de glitchs/bugs que ofereçam vantagem por cima de outros jogadores;\n3・Qualquer tipo de preconceito ou discurso de ódio;\n4・Ghost em streamers (caso comprovado, será paassivel de punição);\n5・Qualquer tipo de combatlog (deslogar durante ações de PvP);\n6・Deixar veículos estacionados dentro de Safezones (recomendamos guarda-los dentro de garagens);\n7・Destruir ou arruinar veículos trancados pelo mapa (dentro ou fora de bases);\n8・Permitido um máximo de 2 helicopteros e 2 carros por clan;\nObservação・NÃO se responsabilizamos por eventuais bugs nos mods que ocasionem perdas ou prejuízo, pois não temos controle sobre os mesmos```")

    await ctx.send(file=discord.File("images/divisoria.gif"))
    embedVar = discord.Embed(title="**Regras de construção**", description="É necessário atentar-se:",colour=discord.Colour.red())
    embedVar.set_footer(text="As regras estão sujeitas a mudanças. Fique atento!")
    await ctx.send(embed=embedVar)
    await ctx.send("```python\n1・Toda e qualquer construção tem que ser feita a mais de 1000 metros das safe-zones;\n2・Toda e qualquer construção tem que ser feita a mais de 800 metros de Black Markets, caçadores e áreas militares;\n3・Toda base precisa ser raidavel;\n4・É permitido apenas uma FOB por clan;\n5・O tamanho máximo de qualquer base é de 125 cubos (exemplo: 5 x 5 x 5);\n6・É estritamente proibido bugar barracas;\n7・Máximo de 10 codelocks por base; Máximo de 3 codelocks por FOB; Sem limite de barracas com codelock (porém, podem ser destruidas em dia de raid);\n8・Todas as paredes devem ter um distanciamento mínimo de 1 jogador;\n9・Proibido colocar portas encostadas uma nas outras;\n10・Toda base precisa ter física básica (recomendamos o uso de pilares e fundação para isso);\n10.1・Toda base que não possua física, será deletada e os itens não serão reembolsados;\n11・Não é permitido colocar codelocks em cofres, armários e etc. (Caso visto, será deletado e os itens não serão reembolsados)```")

    await ctx.send(file=discord.File("images/divisoria.gif"))
    embedVar = discord.Embed(title="**Regras de raid**", description="É necessário atentar-se:",colour=discord.Colour.red())
    embedVar.set_footer(text="As regras estão sujeitas a mudanças. Fique atento!")
    await ctx.send(embed=embedVar)
    await ctx.send('```python\nOs horários de raid são:\n・Sexta-feira: Das 18:00 às 23:59\n・Sábado: Das 14:00 às 00:59\n・Domingo: Das 18:00 às 23:59\n1・O Raid é permitido apenas com C4 em portas, portões e floor/roof hatches;\n2・Não é permitido uso de helicopteros para raidar (não pouse em cima de base alheia);\n3・Toda ação do Raid deve ser gravada e enviada num prazo de 48 horas;\n4・O jogador ou clan que sofreu o Raid, tem até 5 dias para fechar a base. Caso contrário, a base será deletada;\n5・O raid por falha de construção é proibido;\n6・Não é permitido construir enquanto estiver ocorrendo o Raid;\n7・Técnica do pézinho permitido apenas após explodir o floor/roof hatch para acessar o próximo andar;\n8・Proibido usar itens flutuantes ou bugs para raidar;\n9・Proibido empilhar itens para raidar (Ex: empilhar portas de veiculos, barris...);\n10・Proibido utilização de bugs de textura ou qualquer outro tipo de bug;\n11・Barracas com codelock fora de bases podem ser destruidas 24/7 sem a necessidade de gravação;\n12・Na presença de um administrador, caso ocorra um bug ou invalidade durante o raid, ele será interrompido imediatamente;\n13・Proibido destruir veiculos/helicopteros dentro de bases;\n14・Proibido mandar localização de bases abertas em quaisqueres meios de comunicação;\n15・Não é permitido vandalismo por nenhuma das partes;\n16・Quem está sendo raidado não poderá destruir os itens durante o raid.```')
    await ctx.send("```python\nConstrução Tier 1: uma C4 branca OU uma C4 verde\nConstrução Tier 2: duas C4 brancas OU duas C4 verdes\nConstrução Tier 3: quatro C4 verdes.```")
    return

@bot.command(name="regras_2")
async def regras_2(ctx):
    check_1 = discord.utils.get(ctx.guild.roles, id=947242380040478761)

    if check_1 in ctx.author.roles:
        pass
    else:
        warning = await ctx.reply("Você não tem permissão para usar isso!")
        time.sleep(8)
        await warning.delete()
        await ctx.message.delete()
        return

    await ctx.message.delete()

    await ctx.send(file=discord.File("images/regras_namalsk.png"))
    embedVar = discord.Embed(title="**Regras do servidor**", description="É estritamente proibido, dentro das dependências do State of War:",colour=discord.Colour.red())
    embedVar.set_footer(text="As regras estão sujeitas a mudanças. Fique atento!")
    await ctx.send(embed=embedVar)
    await ctx.send("```python\n1・Uso de programas externos que possam oferecer vantagem dentro do jogo (hacks/sem mato);\n2・Uso de glitchs/bugs que ofereçam vantagem por cima de outros jogadores;\n3・Qualquer tipo de preconceito ou discurso de ódio;\n4・Ghost em streamers (caso comprovado, será paassivel de punição);\n5・Qualquer tipo de combatlog (deslogar durante ações de PvP);\n6・Deixar veículos estacionados dentro de Safezones (recomendamos guarda-los dentro de garagens);\n7・Destruir ou arruinar veículos trancados pelo mapa (dentro ou fora de bases);\n8・Permitido um máximo de 2 helicopteros e 2 carros por clan;\nObservação・NÃO se responsabilizamos por eventuais bugs nos mods que ocasionem perdas ou prejuízo, pois não temos controle sobre os mesmos```")

    await ctx.send(file=discord.File("images/divisoria.gif"))
    embedVar = discord.Embed(title="**Regras de construção**", description="É necessário atentar-se:",colour=discord.Colour.red())
    embedVar.set_footer(text="As regras estão sujeitas a mudanças. Fique atento!")
    await ctx.send(embed=embedVar)
    await ctx.send("```python\n1・Toda e qualquer construção tem que ser feita a mais de 1000 metros das safe-zones;\n2・Toda e qualquer construção tem que ser feita a mais de 800 metros de Black Markets, áreas militares;\n2.1・Para áreas militares pequenas e Checkpoints, a distância mínima está reduzida para 300 metros.\n3・Toda base precisa ser raidavel;\n4・É permitido apenas uma FOB por clan;\n5・O tamanho máximo de qualquer base é de 125 cubos (exemplo: 5 x 5 x 5);\n6・É estritamente proibido bugar barracas;\n7・Máximo de 10 codelocks por base; Máximo de 3 codelocks por FOB; Sem limite de barracas com codelock (porém, podem ser destruidas em dia de raid);\n8・Todas as paredes devem ter um distanciamento mínimo de 1 jogador;\n9・Proibido colocar portas encostadas uma nas outras;\n10・Toda base precisa ter física básica (recomendamos o uso de pilares e fundação para isso);\n10.1・Toda base que não possua física, será deletada e os itens não serão reembolsados;\n11・Não é permitido colocar codelocks em cofres, armários e etc. (Caso visto, será deletado e os itens não serão reembolsados)```")

    await ctx.send(file=discord.File("images/divisoria.gif"))
    embedVar = discord.Embed(title="**Regras de raid**", description="É necessário atentar-se:",colour=discord.Colour.red())
    embedVar.set_footer(text="As regras estão sujeitas a mudanças. Fique atento!")
    await ctx.send(embed=embedVar)
    await ctx.send('```python\nOs horários de raid são:\n・Sexta-feira: Das 18:00 às 23:59\n・Sábado: Das 14:00 às 00:59\n・Domingo: Das 18:00 às 23:59\n1・O Raid é permitido apenas com C4 em portas, portões e floor/roof hatches;\n2・Não é permitido uso de helicopteros para raidar (não pouse em cima de base alheia);\n3・Toda ação do Raid deve ser gravada e enviada num prazo de 48 horas;\n4・O jogador ou clan que sofreu o Raid, tem até 5 dias para fechar a base. Caso contrário, a base será deletada;\n5・O raid por falha de construção é proibido;\n6・Não é permitido construir enquanto estiver ocorrendo o Raid;\n7・Técnica do pézinho permitido apenas após explodir o floor/roof hatch para acessar o próximo andar;\n8・Proibido usar itens flutuantes ou bugs para raidar;\n9・Proibido empilhar itens para raidar (Ex: empilhar portas de veiculos, barris...);\n10・Proibido utilização de bugs de textura ou qualquer outro tipo de bug;\n11・Barracas com codelock fora de bases podem ser destruidas 24/7 sem a necessidade de gravação;\n12・Na presença de um administrador, caso ocorra um bug ou invalidade durante o raid, ele será interrompido imediatamente;\n13・Proibido destruir veiculos/helicopteros dentro de bases;\n14・Proibido mandar localização de bases abertas em quaisqueres meios de comunicação;\n15・Não é permitido vandalismo por nenhuma das partes;\n16・Quem está sendo raidado não poderá destruir os itens durante o raid.```')
    await ctx.send("```python\nConstrução Tier 1: uma C4 branca OU uma C4 verde\nConstrução Tier 2: duas C4 brancas OU duas C4 verdes\nConstrução Tier 3: quatro C4 verdes.```")
    return 

@bot.command(name="enquete")
async def enquete(ctx, *,conteudotitulo):
    check_1 = discord.utils.get(ctx.guild.roles, id=947242380040478761)

    if check_1 in ctx.author.roles:
        pass
    else:
        warning = await ctx.reply("Você não tem permissão para usar isso!")
        time.sleep(8)
        await warning.delete()
        await ctx.message.delete()
        return

    conteudotitulo = conteudotitulo.split("/")
    enquete_titulo = conteudotitulo[0]
    enquete_conteudo = conteudotitulo[1]

    channel = bot.get_channel(960031449199829052)

    cor = discord.Colour.random()

    embedVar = discord.Embed(title=f"🗳️ {enquete_titulo}", description=f"{enquete_conteudo}", colour = cor)
    enquete = await channel.send(embed = embedVar)
    await enquete.add_reaction("✅")
    await enquete.add_reaction("❌")
    return

@bot.command(name="code")  ##ARRUMAR O CÓDIGO
async def code(ctx , code):

    if str(datetime.datetime.now().hour) == '8':
        warning = await ctx.reply("Estamos atualizando o nosso banco de dados. Por favor, tente novamente mais tarde\n_Esse processo dura das 05:00 até às 06:00_\nCaso você ache que isso seja um erro, entre em contato com o DEV.")
        time.sleep(10)
        await warning.delete()
        await ctx.message.delete()
        return

    print(f'====================={ctx.author.name} ACABOU DE INICIAR UMA VERIFICAÇÃO=====================')

    authorid = ctx.author.id
    authorid = "`"+str(authorid)+"`"

    async def get_info(codex):
        code = codex
        API_ENDPOINT = 'https://discord.com/api/v8'
        CLIENT_ID = '912310660669521921'
        CLIENT_SECRET = 'TJBNOzRyXWWqlRxCLkLgxDOo2147TLe0'
        REDIRECT_URI = 'https://stateofwar.xyz/verificacao'
        def exchange_code(code):
          data = {
            'client_id': CLIENT_ID,
            'client_secret': CLIENT_SECRET,
            'grant_type': 'authorization_code',
            'code': code,
            'redirect_uri': REDIRECT_URI
          }
          headers = {
            'Content-Type': 'application/x-www-form-urlencoded'
          }
          r = requests.post('%s/oauth2/token' % API_ENDPOINT, data=data, headers=headers)
          r.raise_for_status()
          return r.json()
        def get_user_json(acess_token):
          url = f"{API_ENDPOINT}/users/@me/connections"
          headers = {
            "Authorization": "Bearer {}".format(acess_token)
          }
          user_object = requests.get(url=url, headers=headers)
          user_json = user_object.json()
          return user_json
        code = exchange_code(code)["access_token"]
        user_json = get_user_json(code)
        user = str(user_json).split('}')

        n = 0

        for section in user:
            if ("steam") in section:
                data = section
                n += 1
        
        if n == 0:
            embedVar = discord.Embed(title="⚠️ Steam não encontrada nas suas conexões do Discord.", description="Você precisa vincular a sua Steam a sua conta do Discord para executar a conexão.\nEsse código foi DESCARTADO. Após vincular a sua conta Steam ao seu Discord, gere um novo através do mesmo link e tente novamente.", colour=discord.Colour.random())
            embedVar.set_footer(text="Caso tenha alguma dúvida, entre em contato com o DEV.")
            warning = await ctx.reply(embed=embedVar)
            time.sleep(8)
            await ctx.message.delete()
            await warning.delete()
            return

        data = data[20:]
        data = data.replace("{","")
        data = data.replace("'","")
        data = data.split(",")
        id, nick, verificed = data[0], data[1], data[5]
        id = id.split(':')
        nick = nick.split(':')
        verificed = verificed.split(':')
        steam_id = id[1]
        nick = nick[1]
        verificed = verificed[1]
        steam_id = steam_id[1:]
        nick = nick[1:]
        return steam_id, nick, verificed
    
    data = await get_info(code)

    print(f'OS DADOS FORAM ENCONTRADOS: SEGUEM ABAIXO')
    print(data)

    id = data[0]
    id = "`"+id+"`"
    
    file = "kds_chernarus.xlsx"
    wb = openpyxl.load_workbook(filename=file)
    ws = wb.worksheets[0]
    wsr = wb.active

    for row in wsr.iter_rows(wsr.min_row, wsr.max_row):
        for cell in row:
            if cell.value == None:
                continue
            if str(cell.value) == "":
                continue
            elif str(cell.value) == str(authorid):
                embedVar = discord.Embed(title="⚠️Error 403", description="Já existe um registro com esse DiscordID no nosso banco de dados.", colour=discord.Colour.random())
                embedVar.set_footer(text="Caso tenha alguma dúvida, entre em contato com o DEV.")
                warning = await ctx.reply(embed=embedVar)
                time.sleep(8)
                await ctx.message.delete()
                await warning.delete()
                print("====================FIM - BARRADO EM CHERNARUS - ID DISCORD ====================")
                return
            elif str(cell.value) == str(id):
                embedVar = discord.Embed(title="⚠️Error 403", description="Já existe um registro com esse SteamID no nosso banco de dados.", colour=discord.Colour.random())
                embedVar.set_footer(text="Caso tenha alguma dúvida, entre em contato com o DEV.")
                warning = await ctx.reply(embed=embedVar)
                time.sleep(8)
                await ctx.message.delete()
                await warning.delete()
                print("====================FIM - BARRADO EM CHERNAURS - ID STEAM ====================")
                return


    ws[f'{get_column_letter(ws.min_column + 0)}{(ws.max_row + 1)}'] = authorid #ID do Discord
    ws[f'{get_column_letter(ws.min_column + 1)}{(ws.max_row)}'] = id #ID da Steam
    ws[f'{get_column_letter(ws.min_column + 2)}{(ws.max_row)}'] = "0" #Número de vezes que morreu
    ws[f'{get_column_letter(ws.min_column + 3)}{(ws.max_row)}'] = "0" #Número de vezes que matou
    ws[f'{get_column_letter(ws.min_column + 4)}{(ws.max_row)}'] = "0" #Maior distância de kill
    wb.save(file)
    
    print('PASSOU POR CHERNARUS')

    file = "kds_namalsk.xlsx"
    wb = openpyxl.load_workbook(filename=file)
    ws = wb.worksheets[0]
    wsr = wb.active

    for row in wsr.iter_rows(wsr.min_row, wsr.max_row):
        for cell in row:
            if cell.value == None:
                continue
            if str(cell.value) == "":
                continue
            elif str(cell.value) == str(authorid):
                embedVar = discord.Embed(title="⚠️Error 403", description="Já existe um registro com esse DiscordID no nosso banco de dados.", colour=discord.Colour.random())
                embedVar.set_footer(text="Caso tenha alguma dúvida, entre em contato com o DEV.")
                warning = await ctx.reply(embed=embedVar)
                time.sleep(8)
                await ctx.message.delete()
                await warning.delete()
                print("====================FIM - BARRADO EM NAMALSK - ID DISCORD====================")
                return
            elif str(cell.value) == str(id):
                embedVar = discord.Embed(title="⚠️Error 403", description="Já existe um registro com esse SteamID no nosso banco de dados.", colour=discord.Colour.random())
                embedVar.set_footer(text="Caso tenha alguma dúvida, entre em contato com o DEV.")
                warning = await ctx.reply(embed=embedVar)
                time.sleep(8)
                await ctx.message.delete()
                await warning.delete()
                print("====================FIM - BARRADO EM NAMALSK - ID STEAM====================")
                return

    ws[f'{get_column_letter(ws.min_column + 0)}{(ws.max_row + 1)}'] = authorid #ID do Discord
    ws[f'{get_column_letter(ws.min_column + 1)}{(ws.max_row)}'] = id #ID da Steam
    ws[f'{get_column_letter(ws.min_column + 2)}{(ws.max_row)}'] = "0" #Número de vezes que morreu
    ws[f'{get_column_letter(ws.min_column + 3)}{(ws.max_row)}'] = "0" #Número de vezes que matou
    ws[f'{get_column_letter(ws.min_column + 4)}{(ws.max_row)}'] = "0" #Maior distância de kill
    wb.save(file)

    print('PASSOU POR NAMALSK')

    file = "banco de dados.xlsx"
    wb = openpyxl.load_workbook(filename=file)
    ws = wb.worksheets[0]
    wsr = wb.active

    for row in wsr.iter_rows(wsr.min_row, wsr.max_row):
        for cell in row:
            if cell.value == authorid:
                ws[f'{get_column_letter(cell.column + 2)}{cell.row}'] = id
                ws[f'{get_column_letter(cell.column + 3)}{cell.row}'] = data[1]
                ws[f'{get_column_letter(cell.column + 4)}{cell.row}'] = data[2]
                wb.save(file)

    print('PASSOU PELO BANCO DE DADOS OFICIAL')

    embedVar = discord.Embed(title="✅ Verificado com sucesso", description = "A sua conta Steam está vinculada com o nosso sistema e o nosso sitema de KDr e kills está funcionando corretamente com o seu perfil!.")
    embedVar.set_footer(text="➝ Se precisar de ajuda com o seu registro, entre em contato com o DEV.")
    warning = await ctx.reply(embed=embedVar)

    time.sleep(8)
    await ctx.message.delete()
    await warning.delete()

    print('==================FIM=======================')

@bot.command(name="ficha")
async def ficha(ctx, member:discord.Member):

    if str(datetime.datetime.now().hour) == '8':
        warning = await ctx.reply("Estamos atualizando o nosso banco de dados. Por favor, tente novamente mais tarde\n_Esse processo dura das 05:00 até as 06:00_\nCaso você ache que isso seja um erro, entre em contato com o DEV.")
        time.sleep(10)
        await warning.delete()
        await ctx.message.delete()
        return

    check_1 = discord.utils.get(ctx.guild.roles, id=947242380040478761) #ID DE ADMINISTRADOR SUPREMO

    file = "banco de dados.xlsx"
    wb = openpyxl.load_workbook(filename=file)
    ws = wb.worksheets[0]
    wsr = wb.active

    if check_1 in ctx.author.roles:
        pass
    else:
        warning = await ctx.reply("Você não tem permissão para usar isso!")
        time.sleep(8)
        await warning.delete()
        await ctx.message.delete()
        return

    member_id = member.id
    member_id = str(member_id)
    member_id = "`"+member_id+"`"
    n = 0

    for row in wsr.iter_rows(wsr.min_row, wsr.max_row):
        for cell in row:
            if cell.value == member_id:
                discord_id = ws.cell(row=cell.row, column=1).value
                discord_nick = ws.cell(row=cell.row, column=2).value
                steam_id = ws.cell(row=cell.row, column=3).value
                steam_nick = ws.cell(row=cell.row, column=4).value
                steam_verify = ws.cell(row=cell.row, column=5).value
                rules_broken = ws.cell(row=cell.row, column=6).value
                adv = ws.cell(row=cell.row, column=7).value
                bans = ws.cell(row=cell.row, column=8).value
                clan = ws.cell(row=cell.row, column=9).value
                cars_seguro = ws.cell(row=cell.row, column=10).value
                cars_seguro_days = ws.cell(row=cell.row, column=11).value
                cars_seguro_remain = ws.cell(row=cell.row, column=12).value
                heli_seguro = ws.cell(row=cell.row, column=13).value
                heli_seguro_days = ws.cell(row=cell.row, column=14).value
                heli_seguro_remain = ws.cell(row=cell.row, column=15).value
                construcao = ws.cell(row=cell.row, column=16).value
                construcao_days = ws.cell(row=cell.row, column=17).value
                fila_prioritaria = ws.cell(row=cell.row, column=18).value
                fila_prioritaria_days = ws.cell(row=cell.row, column=19).value

                n += 1
            else:
                n += 0
    if n == 0:
        embedVar = discord.Embed(title="⚠️Error 404", description="Não foi encontrado um registro do usuário em questão!")
        embedVar.set_footer(text="Caso você ache que isso é um engano, contate o DEV!")
        await ctx.send(embed = embedVar)
        return  

    embedVar = discord.Embed(title="📝 Ficha do usuário", description="📝 Segue a ficha detalhada do usuário abaixo\n", colour = discord.Colour.random())
    embedVar.add_field(name="<:steam:958580268581662770> Steam *", value=f"SteamID: {steam_id}\nSteam nick: `{steam_nick}` **\nVerificado: `{steam_verify}`")
    embedVar.add_field(name=":shield: Clã", value=f"{clan}")
    embedVar.add_field(name=":card_box: Histórico", value=f"Infrações: `{rules_broken}`\nAdvertências: `{adv}`\nBanimentos: `{bans}`")
    embedVar.add_field(name="<:discord:958830929810440284> Discord", value=f"Nick: `{discord_nick}`\nID: {discord_id}")
    embedVar.add_field(name=":blue_car: Seguro de carros", value=f"Possui: `{cars_seguro}`\nQuantos restantes: `{cars_seguro_remain}`\nQuantos dias restantes: `{cars_seguro_days}`")
    embedVar.add_field(name=":helicopter: Seguro helicopteros", value=f"Possui: `{heli_seguro}`\nQuantos restantes: `{heli_seguro_remain}`\nQuantos dias restantes: `{heli_seguro_days}`")
    embedVar.add_field(name=":house: Construção", value=f"Possui: `{construcao}`\nDias restantes: `{construcao_days}`")
    embedVar.add_field(name="<:vip:958581319418384395> Fila prioritaria", value=f"Possui: `{fila_prioritaria}`\nDias restantes: `{fila_prioritaria_days}`")
    embedVar.add_field(name="ㅤ", value="ㅤ")
    embedVar.set_footer(text="*Caso as inforamções estejam como 'None', significa que o discord não foi vinculado ao nosso sistema\n** O nick pode ser modificado na Steam e não atualizado no nosso banco de dados.\n➝ Caso tenha alguma dúvida, entre em contato com o DEV.")
    embedVar.set_author(name=member.name, icon_url=member.avatar_url)
    url_antes = f"https://www.steamidfinder.com/lookup/{steam_id}/"

    await ctx.reply(embed=embedVar, components=[Button(label="Mais informações", style = ButtonStyle.URL ,url=url_antes)])
    return

@bot.command(name="perfil")
async def perfil(ctx):

    if str(datetime.datetime.now().hour) == '8':
        warning = await ctx.reply("Estamos atualizando o nosso banco de dados. Por favor, tente novamente mais tarde\n_Esse processo dura das 05:00 até às 06:00_\nCaso você ache que isso seja um erro, entre em contato com o DEV.")
        time.sleep(10)
        await warning.delete()
        await ctx.message.delete()
        return
    
    file = "banco de dados.xlsx"
    wb = openpyxl.load_workbook(filename=file)
    ws = wb.worksheets[0]
    wsr = wb.active

    member = ctx.author

    member_id = member.id
    member_id = str(member_id)
    member_id = "`"+member_id+"`"
    n = 0

    for row in wsr.iter_rows(wsr.min_row, wsr.max_row):
        for cell in row:
            if cell.value == member_id:
                discord_id = ws.cell(row=cell.row, column=1).value
                discord_nick = ws.cell(row=cell.row, column=2).value
                steam_id = ws.cell(row=cell.row, column=3).value
                steam_nick = ws.cell(row=cell.row, column=4).value
                steam_verify = ws.cell(row=cell.row, column=5).value
                rules_broken = ws.cell(row=cell.row, column=6).value
                adv = ws.cell(row=cell.row, column=7).value
                bans = ws.cell(row=cell.row, column=8).value
                clan = ws.cell(row=cell.row, column=9).value
                cars_seguro = ws.cell(row=cell.row, column=10).value
                cars_seguro_days = ws.cell(row=cell.row, column=11).value
                cars_seguro_remain = ws.cell(row=cell.row, column=12).value
                heli_seguro = ws.cell(row=cell.row, column=13).value
                heli_seguro_days = ws.cell(row=cell.row, column=14).value
                heli_seguro_remain = ws.cell(row=cell.row, column=15).value
                construcao = ws.cell(row=cell.row, column=16).value
                construcao_days = ws.cell(row=cell.row, column=17).value
                fila_prioritaria = ws.cell(row=cell.row, column=18).value
                fila_prioritaria_days = ws.cell(row=cell.row, column=19).value

                n += 1
            else:
                n += 0
    if n == 0:
        embedVar = discord.Embed(title="⚠️Error 404", description="Não foi encontrado um registro do usuário em questão!")
        embedVar.set_footer(text="Caso você ache que isso é um engano, contate o DEV!")
        await ctx.send(embed = embedVar)
        return  

    embedVar = discord.Embed(title="📝 Ficha do usuário", description="📝 Segue a ficha detalhada do usuário abaixo\n", colour = discord.Colour.random())
    embedVar.add_field(name="<:steam:958580268581662770> Steam *", value=f"SteamID: {steam_id}\nSteam nick: `{steam_nick}` **\nVerificado: `{steam_verify}`")
    embedVar.add_field(name=":shield: Clã", value=f"{clan}")
    embedVar.add_field(name=":card_box: Histórico", value=f"Infrações: `{rules_broken}`\nAdvertências: `{adv}`\nBanimentos: `{bans}`")
    embedVar.add_field(name="<:discord:958830929810440284> Discord", value=f"Nick: `{discord_nick}`\nID: {discord_id}")
    embedVar.add_field(name=":blue_car: Seguro de carros", value=f"Possui: `{cars_seguro}`\nQuantos restantes: `{cars_seguro_remain}`\nQuantos dias restantes: `{cars_seguro_days}`")
    embedVar.add_field(name=":helicopter: Seguro helicopteros", value=f"Possui: `{heli_seguro}`\nQuantos restantes: `{heli_seguro_remain}`\nQuantos dias restantes: `{heli_seguro_days}`")
    embedVar.add_field(name=":house: Construção", value=f"Possui: `{construcao}`\nDias restantes: `{construcao_days}`")
    embedVar.add_field(name="<:vip:958581319418384395> Fila prioritaria", value=f"Possui: `{fila_prioritaria}`\nDias restantes: `{fila_prioritaria_days}`")
    embedVar.add_field(name="ㅤ", value="ㅤ")
    embedVar.set_footer(text="*Caso as inforamções estejam como 'None', significa que o discord não foi vinculado ao nosso sistema\n** O nick pode ser modificado na Steam e não atualizado no nosso banco de dados.\n➝ Caso tenha alguma dúvida, entre em contato com o DEV.")
    embedVar.set_author(name=member.name, icon_url=member.avatar_url)
    url_antes = f"https://www.steamidfinder.com/lookup/{steam_id}/"
    await member.send(embed=embedVar, components=[Button(label="Mais informações", style = ButtonStyle.URL ,url=url_antes)])
    return

@bot.command(name="registrarall")
async def registrarall(ctx):

    if str(datetime.datetime.now().hour) == '8':
        warning = await ctx.send("Estamos atualizando o nosso banco de dados. Por favor, tente novamente mais tarde\n_Esse processo dura das 05:00 até às 06:00_\nCaso você ache que isso seja um erro, entre em contato com o DEV.")
        time.sleep(10)
        await warning.delete()
        await ctx.message.delete()
        return

    check_1 = discord.utils.get(ctx.guild.roles, id=947242380040478761) #ID DE ADMINISTRADOR SUPREMO

    if check_1 in ctx.author.roles:
        pass
    else:
        warning = await ctx.reply("Você não tem permissão para usar isso!")
        time.sleep(8)
        await warning.delete()
        await ctx.message.delete()
        return 

    guild = bot.get_guild(947237264596041728)
    file = "banco de dados.xlsx"
    wb = openpyxl.load_workbook(filename=file)
    ws = wb.worksheets[0]
    wsr = wb.active
    for member in guild.members:
        id = member.id
        id = str(id)
        id = "`"+id+"`"
        name = member.name
        discriminator = member.discriminator
        nickname = name + "#" + discriminator
        ws[f'{get_column_letter(ws.min_column + 0)}{(ws.max_row + 1)}'] = id
        ws[f'{get_column_letter(ws.min_column + 1)}{(ws.max_row)}'] = nickname
        ws[f'{get_column_letter(ws.min_column + 5)}{(ws.max_row)}'] = "0" #INFRAÇÕES
        ws[f'{get_column_letter(ws.min_column + 6)}{(ws.max_row)}'] = "0" #ADVERTÊNCIAS
        ws[f'{get_column_letter(ws.min_column + 7)}{(ws.max_row)}'] = "0" #BANIMENTOS
        ws[f'{get_column_letter(ws.min_column + 8)}{(ws.max_row)}'] = "Não participante" #CLÃ
        ws[f'{get_column_letter(ws.min_column + 9)}{(ws.max_row)}'] = "❌" #SEGURO DE VEICULOS
        ws[f'{get_column_letter(ws.min_column + 10)}{(ws.max_row)}'] = "0" #QUANTOS DIAS RESTANTES
        ws[f'{get_column_letter(ws.min_column + 11)}{(ws.max_row)}'] = "Indefinido" #QUANTOS SEGUROS RESTANTES DE VEICULOS
        ws[f'{get_column_letter(ws.min_column + 12)}{(ws.max_row)}'] = "❌" #SEGURO DE HELICOPTEROS
        ws[f'{get_column_letter(ws.min_column + 13)}{(ws.max_row)}'] = "0" #QUANTOS DIAS RESTANTES
        ws[f'{get_column_letter(ws.min_column + 14)}{(ws.max_row)}'] = "Indefinido" #QUANTOS SEGUROS RESTANTES DE HELICOPTEROS
        ws[f'{get_column_letter(ws.min_column + 15)}{(ws.max_row)}'] = "❌" #POSSUI ALGUMA CONSTRUÇÃO
        ws[f'{get_column_letter(ws.min_column + 16)}{(ws.max_row)}'] = "0" #QUANTOS DIAS RESTANTES DA CONSTRUÇÃO
        ws[f'{get_column_letter(ws.min_column + 17)}{(ws.max_row)}'] = "❌" #POSSUI FILA PRIORITARIA
        ws[f'{get_column_letter(ws.min_column + 18)}{(ws.max_row)}'] = "0" #QUANTOS DIAS RESTANTES 
        wb.save(file)
    print('done')
    return

@bot.command(name="vipadd")
async def vipadd(ctx, member:discord.Member):

    if str(datetime.datetime.now().hour) == '8':
        warning = await ctx.send("Estamos atualizando o nosso banco de dados. Por favor, tente novamente mais tarde\n_Esse processo dura das 05:00 até às 06:00_\nCaso você ache que isso seja um erro, entre em contato com o DEV.")
        time.sleep(10)
        await warning.delete()
        await ctx.message.delete()
        return

    check_1 = discord.utils.get(ctx.guild.roles, id=947242380040478761) #ID DE ADMINISTRADOR SUPREMO

    if check_1 in ctx.author.roles:
        pass
    else:
        warning = await ctx.reply("Você não tem permissão para usar isso!")
        time.sleep(8)
        await warning.delete()
        await ctx.message.delete()
        return 

    file = "banco de dados.xlsx"
    wb = openpyxl.load_workbook(filename=file)
    ws = wb.worksheets[0]
    wsr = wb.active

    member_id = member.id
    member_id = str(member_id)
    member_id = "`"+member_id+"`"
    n = 0


    embedVar = discord.Embed(title=f"<:vip:958581319418384395> Adicionar VIP para **{member.name}**", description="Selecione o VIP que você deseja adicionar.", colour = discord.Colour.random())
    embedVar.add_field(name="Observações:", value="・Por questões de segurança, apenas o autor do comando poderá adicionar o VIP.\n・O comando deixará de funcionar em 10 segundos.\n・O tempo padrão de qualquer doação é de 30 dias.\n・Ignore a mensagem _'interação falhou'_, é um bug do Discord.")
    embedVar.set_author(name=member.name, icon_url=member.avatar_url)
    embedVar.set_footer(text="➝ Caso tenha alguma dúvida, entre em contato com o DEV.")
    
    await ctx.send(embed=embedVar, components=[Select(
        placeholder="Adicionar tipo de doação",
        options=[
            SelectOption(label="Fila prioritária", value="filaprioritaria", emoji="⭐"),
            SelectOption(label="Seguro de veículos", value="seguro_de_veiculos", emoji="🚘"),
            SelectOption(label="Seguro de helicopteros", value="seguro_de_heli", emoji="🚁"),
            SelectOption(label="Bunker menor", value="bunker_menor", emoji="🏠"),
            SelectOption(label="Bunker maior", value="bunker_maior", emoji="🏠"),
            SelectOption(label="Prédio", value="predio", emoji="🏠")

        ],
        custom_id='VIP_List'
    )])

    interaction = await bot.wait_for('select_option', check=lambda inter: inter.custom_id == 'VIP_List' and inter.user == ctx.author, timeout = 10)
    
    if interaction.values[0] == "filaprioritaria":
        for row in wsr.iter_rows(wsr.min_row, wsr.max_row):
            for cell in row:
                if cell.value == member_id:
                    ws[f'{get_column_letter(cell.column + 17)}{(cell.row)}'] = "✅" #POSSUI FILA PRIORITARIA
                    ws[f'{get_column_letter(cell.column + 18)}{(cell.row)}'] = "31" #TEMPO RESTANTE DA VILA PRIORITARIA
                    wb.save(file)
                    n += 1
                    id = "fila prioritária"
                else:
                    n += 0
    elif interaction.values[0] == "seguro_de_veiculos":
        for row in wsr.iter_rows(wsr.min_row, wsr.max_row):
            for cell in row:
                if cell.value == member_id:
                    ws[f'{get_column_letter(cell.column + 9)}{(cell.row)}'] = "✅" #POSSUI FILA PRIORITARIA
                    ws[f'{get_column_letter(cell.column + 10)}{(cell.row)}'] = "31" #TEMPO RESTANTE DA VILA PRIORITARIA
                    ws[f'{get_column_letter(cell.column + 11)}{(cell.row)}'] = "Indefinido" #SEGUROS RESTANTES
                    wb.save(file)
                    n += 1
                    id = "seguro de veículos"
                else:
                    n += 0
    elif interaction.values[0] == "seguro_de_heli":
        for row in wsr.iter_rows(wsr.min_row, wsr.max_row):
            for cell in row:
                if cell.value == member_id:
                    ws[f'{get_column_letter(cell.column + 12)}{(cell.row)}'] = "✅" #POSSUI FILA PRIORITARIA
                    ws[f'{get_column_letter(cell.column + 13)}{(cell.row)}'] = "31" #TEMPO RESTANTE DA VILA PRIORITARIA
                    ws[f'{get_column_letter(cell.column + 14)}{(cell.row)}'] = "Indefinido" #SEGUROS RESTANTES
                    wb.save(file)
                    n += 1
                    id = "seguro de helicopteros"
                else:
                    n += 0
    elif interaction.values[0] == "bunker_menor":
        for row in wsr.iter_rows(wsr.min_row, wsr.max_row):
            for cell in row:
                if cell.value == member_id:
                    ws[f'{get_column_letter(cell.column + 15)}{(cell.row)}'] = "Bunker menor" #CONSTRUÇÃO EM POSSE
                    ws[f'{get_column_letter(cell.column + 16)}{(cell.row)}'] = "31" #TEMPO DE POSSE DA CONSTRUÇÃO
                    wb.save(file)
                    n += 1
                    id = "bunker menor"
                else:
                    n += 0
    elif interaction.values[0] == "bunker_maior":
        for row in wsr.iter_rows(wsr.min_row, wsr.max_row):
            for cell in row:
                if cell.value == member_id:
                    ws[f'{get_column_letter(cell.column + 15)}{(cell.row)}'] = "Bunker maior" #CONSTRUÇÃO EM POSSE
                    ws[f'{get_column_letter(cell.column + 16)}{(cell.row)}'] = "31" #TEMPO DE POSSE DA CONSTRUÇÃO
                    wb.save(file)
                    n += 1
                    id = "bunker maior"
                else:
                    n += 0
    elif interaction.values[0] == "predio":
        for row in wsr.iter_rows(wsr.min_row, wsr.max_row):
            for cell in row:
                if cell.value == member_id:
                    ws[f'{get_column_letter(cell.column + 15)}{(cell.row)}'] = "Prédio" #CONSTRUÇÃO EM POSSE
                    ws[f'{get_column_letter(cell.column + 16)}{(cell.row)}'] = "31" #TEMPO DE POSSE DA CONSTRUÇÃO
                    wb.save(file)
                    n += 1
                    id = "prédio"
                else:
                    n += 0
    
    if n == 0:
        embedVar = discord.Embed(title="⚠️Error 404", description="Não foi encontrado um registro do usuário em questão!")
        embedVar.set_footer(text="Caso você ache que isso é um engano, contate o DEV!")
        await ctx.send(embed = embedVar)
        return

    confirmacao = discord.Embed(title="✅ Adicionado com sucesso!", description=f"Foi adicionado a vantagem `{id}` com sucesso ao jogador {member.name}!", colour=discord.Colour.random())
    confirmacao.set_footer(text="Caso tenha alguma dúvida, entre em contato com o DEV.")

    await ctx.send(embed=confirmacao)

@bot.command(name="vipmod")
async def vipmod(ctx, member:discord.Member):

    if str(datetime.datetime.now().hour) == '8':
        warning = await ctx.send("Estamos atualizando o nosso banco de dados. Por favor, tente novamente mais tarde\n_Esse processo dura das 05:00 até ás 06:00_\nCaso você ache que isso seja um erro, entre em contato com o DEV.")
        time.sleep(10)
        await warning.delete()
        await ctx.message.delete()
        return

    check_1 = discord.utils.get(ctx.guild.roles, id=947242380040478761)

    if check_1 in ctx.author.roles:
        pass
    else:
        warning = await ctx.reply("Você não tem permissão para usar esse comando!")
        time.sleep(8)
        await warning.delete()
        await ctx.message.delete()
        return

    await ctx.message.delete()

    def check_message(msg):
        return msg.author == ctx.author

    member_id = "`" + str(member.id) + "`"

    file = "banco de dados.xlsx"
    wb = openpyxl.load_workbook(filename=file)
    ws = wb.worksheets[0]
    wsr = wb.active

    n = 0

    for row in wsr.iter_rows(wsr.min_row, wsr.max_row):
        for cell in row:
            if cell.value == member_id:
                n += 1
                embedVar = discord.Embed(title=f"<:vip:958581319418384395> Modificação de benefícios do {member.name}", description="Por favor, selecione o benefício que deseja modificar.", colour=discord.Colour.random())
                embedVar.add_field(name="Observações:", value="・Por questões de segurança, apenas o autor do comando poderá modificar os benefícios.\n・O comando deixará de funcionar em 10 segundos.\n・Ignore a mensagem _'interação falhou'_, é um bug do Discord.")
                embedVar.set_author(name=member.name, icon_url=member.avatar_url)
                embedVar.set_footer(text="Caso tenha algum problema com esse comando, contate o DEV.")
                mensagem_select = await ctx.send(embed=embedVar, components=[Select(
        placeholder="Adicionar tipo de doação",
        options=[
            SelectOption(label="Fila prioritária (dias)", value="filaprioritaria", emoji="⭐"),
            SelectOption(label="Seguro de veículos (dias)", value="seguro_de_veiculos_dias", emoji="🚘"),
            SelectOption(label="Seguro de veículos (quantidade)", value="seguro_de_veiculos_quantidade", emoji="🚘"),
            SelectOption(label="Seguro de helicopteros (dias)", value="seguro_de_heli_dias", emoji="🚁"),
            SelectOption(label="Seguro de helicopteros (quantidade)", value="seguro_de_heli_quantidade", emoji="🚁"),
            SelectOption(label="Tipo de construção", value="tipo_de_construcao", emoji="🏠"),
            SelectOption(label="Dias da construção", value="dias_de_construcao", emoji="🏠")
        ],
        custom_id='MOD_List'
    )])

    interaction = await bot.wait_for('select_option', check=lambda inter: inter.custom_id == 'MOD_List' and inter.user == ctx.author, timeout=10)
    await mensagem_select.delete()


    if interaction.values[0] == "filaprioritaria":
        embedVar = discord.Embed(title=f"Alterar fila prioritária de {member.name}", description=f"Por favor, digite a quantidade de dias que você deseja colocar para a fila prioritária de {member.name}", colour = discord.Colour.random())
        embedVar.add_field(name="Observações:", value="・Por questões de segurança, apenas o autor do comando poderá modificar os benefícios.\n・Caso tenha se equivocado no comando, apenas ignore **POR 15 SEGUNDOS**.\n・Caso você modifique os dias para 0, a fila prioritária será automaticamente desativada no banco de dados.")
        embedVar.set_author(name=member.name, icon_url=member.avatar_url)
        embedVar.set_footer(text="Caso tenha algum problema com esse comando, contate o DEV.")
        mensagem = await ctx.send(embed=embedVar)
        message = await bot.wait_for("message", timeout=15, check=check_message)
        if message.content == '0':
            for row in wsr.iter_rows(wsr.min_row, wsr.max_row):
                for cell in row:
                    if cell.value == member_id:
                        ws[f'{get_column_letter(cell.column + 17)}{cell.row}'] = '❌'
                        ws[f'{get_column_letter(cell.column + 18)}{cell.row}'] = '0'
                        wb.save(file)

        else:
            for row in wsr.iter_rows(wsr.min_row, wsr.max_row):
                for cell in row:
                    if cell.value == member_id:
                        ws[f'{get_column_letter(cell.column + 17)}{cell.row}'] = '✅'
                        ws[f'{get_column_letter(cell.column + 18)}{cell.row}'] = str(message.content)
                        wb.save(file) 

    elif interaction.values[0] == "seguro_de_veiculos_dias":
        embedVar = discord.Embed(title=f"Alterar seguro de veículos de {member.name}", description=f"Por favor, digite a quantidade de dias que você deseja colocar para o seguro de veículos de {member.name}", colour = discord.Colour.random())
        embedVar.add_field(name="Observações:", value="・Por questões de segurança, apenas o autor do comando poderá modificar os benefícios.\n・Caso tenha se equivocado no comando, apenas ignore **POR 15 SEGUNDOS**.\n・Caso você modifique os dias para 0, o seguro de veículos será automaticamente desativado no banco de dados.")
        embedVar.set_author(name=member.name, icon_url=member.avatar_url)
        embedVar.set_footer(text="Caso tenha algum problema com esse comando, contate o DEV.")
        mensagem = await ctx.send(embed=embedVar)
        message = await bot.wait_for("message", timeout=15, check=check_message)
        if message.content == '0':
            for row in wsr.iter_rows(wsr.min_row, wsr.max_row):
                for cell in row:
                    if cell.value == member_id:
                        ws[f'{get_column_letter(cell.column + 9)}{cell.row}'] = '❌'
                        ws[f'{get_column_letter(cell.column + 10)}{cell.row}'] = '0'
                        wb.save(file)
        else:
            for row in wsr.iter_rows(wsr.min_row, wsr.max_row):
                for cell in row:
                    if cell.value == member_id:
                        ws[f'{get_column_letter(cell.column + 9)}{cell.row}'] = '✅'
                        ws[f'{get_column_letter(cell.column + 10)}{cell.row}'] = str(message.content)
                        wb.save(file)
    
    elif interaction.values[0] == "seguro_de_veiculos_quantidade":
        embedVar = discord.Embed(title=f"Alterar seguro de veículos de {member.name}", description=f"Por favor, digite a quantidade de seguros que você deseja colocar para o seguro de veículos de {member.name}", colour = discord.Colour.random())
        embedVar.add_field(name="Observações:", value="・Por questões de segurança, apenas o autor do comando poderá modificar os benefícios.\n・Caso tenha se equivocado no comando, apenas ignore **POR 15 SEGUNDOS**.")
        embedVar.set_author(name=member.name, icon_url=member.avatar_url)
        embedVar.set_footer(text="Caso tenha algum problema com esse comando, contate o DEV.")
        mensagem = await ctx.send(embed=embedVar)
        message = await bot.wait_for("message", timeout=15, check=check_message)
        for row in wsr.iter_rows(wsr.min_row, wsr.max_row):
                for cell in row:
                    if cell.value == member_id:
                        ws[f'{get_column_letter(cell.column + 11)}{cell.row}'] = str(message.content)
                        wb.save(file)

    elif interaction.values[0] == "seguro_de_heli_dias":
        embedVar = discord.Embed(title=f"Alterar o seguro de helicopteros de {member.name}", description=f"Por favor, digite a quantidade de dias que você deseja colocar para o seguro de helicopteros de {member.name}", colour = discord.Colour.random())
        embedVar.add_field(name="Observações:", value="・Por questões de segurança, apenas o autor do comando poderá modificar os benefícios.\n・Caso tenha se equivocado no comando, apenas ignore **POR 15 SEGUNDOS**.\n・Caso você modifique os dias para 0, o seguro de helicopteros será automaticamente desativado do nosso banco de dados.")
        embedVar.set_author(name=member.name, icon_url=member.avatar_url)
        embedVar.set_footer(text="Caso tenha algum problema com esse comando, contate o DEV.")
        mensagem = await ctx.send(embed=embedVar)
        message = await bot.wait_for("message", timeout=15, check=check_message)
        if message.content == '0':
            for row in wsr.iter_rows(wsr.min_row, wsr.max_row):
                for cell in row:
                    if cell.value == member_id:
                        ws[f'{get_column_letter(cell.column + 12)}{cell.row}'] = '❌'
                        ws[f'{get_column_letter(cell.column + 13)}{cell.row}'] = '0'
                        wb.save(file)
        else:
            for row in wsr.iter_rows(wsr.min_row, wsr.max_row):
                for cell in row:
                    if cell.value == member_id:
                        ws[f'{get_column_letter(cell.column + 12)}{cell.row}'] = '✅'
                        ws[f'{get_column_letter(cell.column + 13)}{cell.row}'] = str(message.content)
                        wb.save(file)

    elif interaction.values[0] == "seguro_de_heli_quantidade":
        embedVar = discord.Embed(title=f"Alterar seguro de helicopteros de {member.name}", description=f"Por favor, digite a quantidade de seguros que você deseja colocar para o seguro de helicopteros de {member.name}", colour = discord.Colour.random())
        embedVar.add_field(name="Observações:", value="・Por questões de segurança, apenas o autor do comando poderá modificar os benefícios.\n・Caso tenha se equivocado no comando, apenas ignore **POR 15 SEGUNDOS**.")
        embedVar.set_author(name=member.name, icon_url=member.avatar_url)
        embedVar.set_footer(text="Caso tenha algum problema com esse comando, contate o DEV.")
        mensagem = await ctx.send(embed=embedVar)
        message = await bot.wait_for("message", timeout=15, check=check_message)
        for row in wsr.iter_rows(wsr.min_row, wsr.max_row):
                for cell in row:
                    if cell.value == member_id:
                        ws[f'{get_column_letter(cell.column + 14)}{cell.row}'] = str(message.content)
                        wb.save(file)
    
    elif interaction.values[0] == "tipo_de_construcao":
        embedVar = discord.Embed(title=f"Mudar tipo de construção em posse de {member.name}", description=f"Por favor, selecione a nova construção a ser colocada em posse de {member.mention}", colour=discord.Colour.random())
        embedVar.add_field(name="Observações:", value="・Por questões de segurança, apenas o autor do comando poderá modificar os benefícios.\n・Caso tenha se equivocado no comando, apenas ignore **POR 15 SEGUNDOS**.")
        embedVar.set_author(name=member.name, icon_url=member.avatar_url)
        embedVar.set_footer(text="Caso tenha algum problema com esse comando, contate o DEV.")
        mensagem = await ctx.send(embed=embedVar, components=[Select(
        placeholder="Tipo de construção",
        options=[
            SelectOption(label="Bunker pequeno", value="bunker_pequeno", emoji="🏠"),
            SelectOption(label="Bunker médio", value="bunker_medio", emoji="🏠"),
            SelectOption(label="Prédio", value="predio", emoji="🏠")
        ],
        custom_id='MOD_List_Cons'
        )])

        interaction_ = await bot.wait_for('select_option', check=lambda inter: inter.custom_id == 'MOD_List_Cons' and inter.user == ctx.author)

        if interaction_.values[0] == "bunker_pequeno":
            for row in wsr.iter_rows(wsr.min_row, wsr.max_row):
                for cell in row:
                    if cell.value == member_id:
                        ws[f'{get_column_letter(cell.column + 15)}{cell.row}'] = "Bunker pequeno"
                        wb.save(file)
        
        elif interaction_.values[0] == "bunker_medio":
            for row in wsr.iter_rows(wsr.min_row, wsr.max_row):
                for cell in row:
                    if cell.value == member_id:
                        ws[f'{get_column_letter(cell.column + 15)}{cell.row}'] = "Bunker médio"
                        wb.save(file)
        
        elif interaction_.values[0] == "predio":
            for row in wsr.iter_rows(wsr.min_row, wsr.max_row):
                for cell in row:
                    if cell.value == member_id:
                        ws[f'{get_column_letter(cell.column + 15)}{cell.row}'] = "Prédio"
                        wb.save(file)
        
    elif interaction.values[0] == "dias_de_construcao":
        embedVar = discord.Embed(title=f"Alterar tempo de construção {member.name}", description=f"Por favor, digite a quantidade de dias que você deseja colocar para a construção em posse de {member.name}", colour = discord.Colour.random())
        embedVar.add_field(name="Observações:", value="・Por questões de segurança, apenas o autor do comando poderá modificar os benefícios.\n・Caso tenha se equivocado no comando, apenas ignore **POR 15 SEGUNDOS**.\n・Caso você modifique os dias para 0, a posse da construção será automaticamente desativada do nosso banco de dados.")
        embedVar.set_author(name=member.name, icon_url=member.avatar_url)
        embedVar.set_footer(text="Caso tenha algum problema com esse comando, contate o DEV.")
        mensagem = await ctx.send(embed=embedVar)
        message = await bot.wait_for("message", timeout=15, check=check_message)
        if message.content == '0':
            for row in wsr.iter_rows(wsr.min_row, wsr.max_row):
                for cell in row:
                    if cell.value == member_id:
                        ws[f'{get_column_letter(cell.column + 15)}{cell.row}'] = '❌'
                        ws[f'{get_column_letter(cell.column + 16)}{cell.row}'] = '0'
                        wb.save(file)
        else:
            for row in wsr.iter_rows(wsr.min_row, wsr.max_row):
                for cell in row:
                    if cell.value == member_id:
                        ws[f'{get_column_letter(cell.column + 16)}{cell.row}'] = str(message.content)
                        wb.save(file)

    embedVar = discord.Embed(title="✅ Alterado com sucesso", description=f"Os benefícios de {member.name} foram alterados com sucesso do nosso banco de dados")
    embedVar.set_author(name=member.name, icon_url=member.avatar_url)
    embedVar.set_footer(text="Caso tenha alguma dúvida, entre em contato com o DEV.")

    await mensagem.delete()
    mensagem = await ctx.send(embed=embedVar)
    await message.delete()
    time.sleep(8)
    await mensagem.delete()
    
@bot.command(name="da")
async def doacoesanuncios(ctx):
    check_1 = discord.utils.get(ctx.guild.roles, id=947242380040478761)

    if check_1 in ctx.author.roles:
        pass
    else:
        warning = await ctx.reply("Você não tem permissão para usar isso!")
        time.sleep(8)
        await warning.delete()
        await ctx.message.delete()
        return

    await ctx.message.delete()
    await ctx.send(file=discord.File('images/doacoes.png'))

    embedVar = discord.Embed(title="DOAÇÕES PARA O STATE OF WAR", description="Aprenda aqui como doar para o servidor!", colour = discord.Colour.random())
    embedVar.add_field(name="Porque doar?", value="Doar é uma maneira de ajudar o nosso servidor se manter ativo. Todo dinheiro recebido será investido em melhoras de equipamentos, host e qualidade do servidor.", inline=False)
    embedVar.add_field(name="Qual quantia eu posso doar?", value="Nós aceitamos doações de valores dentro de uma tabela de preços, para melhor administração do dinheiro.", inline=False)
    embedVar.add_field(name="Como posso doar?", value="Para ter mais informações sobre como ajudar nosso servidor, clique no botão abaixo 'Saiba mais!' ou entre em nosso site!")
    embedVar.set_footer(text="➝ Caso o Bot estiver offline, esse comando não vai funcionar!")

    await ctx.send(embed=embedVar, components=[Button(label="Saiba mais!", style=ButtonStyle.green, custom_id="more_info_vip"), Button(label="Site", style=ButtonStyle.URL, url="https://stateofwar.xyz/doacoes")])

@bot.command(name="punir")
async def punir(ctx, member:discord.Member):
    
    if str(datetime.datetime.now().hour) == '8':
        warning = await ctx.reply("Estamos atualizando o nosso banco de dados. Por favor, tente novamente mais tarde\n_Esse processo dura das 05:00 até às 06:00_\nCaso você ache que isso seja um erro, entre em contato com o DEV.")
        time.sleep(10)
        await warning.delete()
        await ctx.message.delete()
        return

    check_1 = discord.utils.get(ctx.guild.roles, id=947242380040478761) #ID DE ADMINISTRADOR SUPREMO

    if check_1 in ctx.author.roles:
        pass
    else:
        warning = await ctx.reply("Você não tem permissão para usar isso!")
        time.sleep(8)
        await warning.delete()
        await ctx.message.delete()
        return

    await ctx.message.delete() 

    member_id = member.id
    member_id = str(member_id)
    member_id = "`"+member_id+"`"
    n = 0

    embedVar = discord.Embed(title=f"🚨 Punir {member.name}", description=f"Selecione o tipo de punição que você deseja adicionar ao jogador {member.name}", colour = discord.Colour.random())
    embedVar.add_field(name="Observações:", value="・Por questões de segurança, apenas o autor do comando poderá adicionar a punição.\n・Caso tenha se equivocado no comando, apenas ignore.\n・Ignore o erro _Esta interação falhou_, isso é apenas um erro do Discord.")
    embedVar.set_footer(text="Caso tenha alguma duvida, entre em contato com o DEV.")
    embedVar.set_author(name=member.name, icon_url=member.avatar_url)

    mensagem = await ctx.send(embed=embedVar, components=[Select(placeholder="Punição", custom_id="punicao", options=[
        SelectOption(label="Infração", value="infracao", emoji="🚨"),
        SelectOption(label="Advertência", value="advertencia", emoji="🚨"),
        SelectOption(label="Banimento", value="banimento", emoji="🚨")
    ])])
    interaction = await bot.wait_for('select_option', check=lambda inter: inter.custom_id == 'punicao' and inter.user == ctx.author)

    file = "banco de dados.xlsx"
    wb = openpyxl.load_workbook(filename=file)
    ws = wb.worksheets[0]
    wsr = wb.active

    if interaction.values[0] == "infracao":
        for row in wsr.iter_rows(wsr.min_row, wsr.max_row):
            for cell in row:
                if cell.value == member_id:
                    antigo = ws.cell(row=cell.row, column=6).value
                    antigo = int(antigo)
                    novo = antigo + 1
                    novo = str(novo)
                    ws[f'{get_column_letter(cell.column + 5)}{(cell.row)}'] = novo
                    wb.save(file)
                    n += 1
                else:
                    n += 0
    
    elif interaction.values[0] == "advertencia":
        for row in wsr.iter_rows(wsr.min_row, wsr.max_row):
            for cell in row:
                if cell.value == member_id:
                    antigo = ws.cell(row=cell.row, column=7).value
                    antigo = int(antigo)
                    novo = antigo + 1
                    novo = str(novo)
                    ws[f'{get_column_letter(cell.column + 6)}{(cell.row)}'] = novo #POSSUI FILA PRIORITARIA
                    wb.save(file)
                    n += 1
                else:
                    n += 0
    
    elif interaction.values[0] == "banimento":
        for row in wsr.iter_rows(wsr.min_row, wsr.max_row):
            for cell in row:
                if cell.value == member_id:
                    antigo = ws.cell(row=cell.row, column=8).value
                    antigo = int(antigo)
                    novo = antigo + 1
                    novo = str(novo)
                    ws[f'{get_column_letter(cell.column + 7)}{(cell.row)}'] = novo #POSSUI FILA PRIORITARIA
                    wb.save(file)
                    n += 1
                else:
                    n += 0

    if n == 0:
        embedVar = discord.Embed(title="⚠️Error 404", description="Não foi encontrado um registro do usuário em questão!")
        embedVar.set_footer(text="Caso você ache que isso é um engano, contate o DEV!")
        await ctx.send(embed = embedVar)
        return
    
    await mensagem.delete()

    embedVar = discord.Embed(title="🚨 Punição concedida", description=f"Foi adicionado com sucesso a punição ao jogador {member.name}", colour=discord.Colour.random())
    embedVar.set_author(name=member.name, icon_url=member.avatar_url)
    embedVar.set_footer(text="Caso tenha alguma dúvida, entre em contato com o DEV.")

    mensagem = await ctx.send(embed=embedVar)
    time.sleep(8)
    await mensagem.delete()
    return

@bot.command(name="clan_criar")
async def clan_criar(ctx, tag, r=0, g=0, b=0):

    if r == 0 or r == '0':
        if g == 0 or g == '0':
            if b == 0 or b == '0':
                no_colour = True
    
    else:
        r,g,b = int(r), int(g), int(b)

    if str(datetime.datetime.now().hour) == '8':
        warning = await ctx.reply("Estamos atualizando o nosso banco de dados. Por favor, tente novamente mais tarde\n_Esse processo dura das 05:00 até ás 06:00_\nCaso você ache que isso seja um erro, entre em contato com o DEV.")
        time.sleep(10)
        await warning.delete()
        await ctx.message.delete()
        return

    id = ctx.author.id
    member_id = "`" + str(id) + "`"

    file = "banco de dados.xlsx"
    wb = openpyxl.load_workbook(filename=file)
    ws = wb.worksheets[0]
    wsr = wb.active
    n = 0

    for row in wsr.iter_cols(min_col=9, max_col=9, min_row=1):
        for cell in row:
            if cell.value == f'{tag}' and cell.value != None:
                warning = await ctx.send("⚠️ Já existe um clan com esse nome!\n⚠️ Por favor, tente outro nome.")
                time.sleep(8)
                await warning.delete()
                return
            elif cell.value == (f"{tag} " + "(Líder)") and cell.value != None:
                warning = await ctx.send("⚠️ Já existe um clan com esse nome!\n⚠️ Por favor, tente outro nome.")
                time.sleep(8)
                await warning.delete()
                return

    for row in wsr.iter_rows(wsr.min_row, wsr.max_row):
        for cell in row:
            if cell.value == member_id:
                clan = ws.cell(row=cell.row, column=9).value
                n += 1
            else:
                n += 0 
                
    if n == 0:
        embedVar = discord.Embed(title="⚠️Error 404", description="Você não foi encontrado em nosso banco de dados!", colour=discord.Colour.random())
        embedVar.set_author(name=ctx.author.name, icon_url=ctx.author.avatar_url)
        embedVar.set_footer(text="Por favor, entre em contato urgentemente com o DEV.")
        await ctx.reply(embed = embedVar)
        return

    if clan != "Não participante":
        embedVar = discord.Embed(title=f"⚠️ Você já pertence ao clan {clan}!", description="Não foi possível criar um novo clan pois você já pertence a um!", colour=discord.Colour.random())
        embedVar.set_footer(text="➝ Se precisar de ajuda, entre em contato com o DEV.")
        embedVar.set_author(name=ctx.author.name, icon_url=ctx.author.avatar_url)
        await ctx.reply(embed = embedVar)
        return
    else:
        pass

    if no_colour:
        role = await ctx.guild.create_role(name=f"{tag}", colour=discord.Colour.random(), hoist = True)
    else:
        role = await ctx.guild.create_role(name=f"{tag}", colour=discord.Colour.from_rgb(r, g, b), hoist=True)

    await ctx.author.add_roles(role)

    for row in wsr.iter_rows(wsr.min_row, wsr.max_row):
        for cell in row:
            if cell.value == member_id:
                ws[f'{get_column_letter(cell.column + 8)}{(cell.row)}'] = f"{tag}" + " (Líder)"
                wb.save(file)
    

    embedVar = discord.Embed(title=f"O clan {tag} foi criado com sucesso!", description=f"O líder do mais novo clan {tag} é o {ctx.author.name}", colour=discord.Colour.random())
    embedVar.set_author(name=ctx.author.name, icon_url=ctx.author.avatar_url)
    embedVar.add_field(name="Observações:", value="・Somente o líder do clan poderá convidar e expulsar membros do clan.\n・Para modificar a TAG ou a cor do clan, entre em contato com um ADM.\n・Usar `!clan_excluir` irá deletar o clan permanentemente.")
    embedVar.set_footer(text="➝ Se precisar de ajuda, entre em contato com o DEV.")
    await ctx.reply(embed=embedVar)
    await role.edit(position=6)
    return

@bot.command(name="clan_sair")
async def clan_sair(ctx):
      
    if str(datetime.datetime.now().hour) == '8':
        warning = await ctx.reply("Estamos atualizando o nosso banco de dados. Por favor, tente novamente mais tarde\n_Esse processo dura das 05:00 até as 06:00_\nCaso você ache que isso seja um erro, entre em contato com o DEV.")
        time.sleep(10)
        await warning.delete()
        await ctx.message.delete()
        return

    id = ctx.author.id
    member_id = "`" + str(id) + "`"

    file = "banco de dados.xlsx"
    wb = openpyxl.load_workbook(filename=file)
    ws = wb.worksheets[0]
    wsr = wb.active
    n = 0

    for row in wsr.iter_rows(wsr.min_row, wsr.max_row):
        for cell in row:
            if cell.value == member_id:
                clan = ws.cell(row=cell.row, column=9).value
                n += 1
            else:
                n += 0

    clan_ = clan

    if clan == "Não participante":
        embedVar = discord.Embed(title="⚠️Error 401", description="Você não participa de um clan", colour=discord.Colour.random())
        embedVar.set_author(name=ctx.author.name, icon_url=ctx.author.avatar_url)
        embedVar.set_footer(text="➝ Se precisar de ajuda, entre em contato com o DEV.")
        await ctx.reply(embed = embedVar)
        return
    
    if "Líder" in clan:
        embedVar = discord.Embed(title="⚠️Error 403", description="Você é líder do seu clan!\nVocê não pode abandonar o seu clan, mas você pode excluí-lo.", colour=discord.Colour.random())
        embedVar.set_author(name=ctx.author.name, icon_url=ctx.author.avatar_url)
        embedVar.set_footer(text="➝ Se precisar de ajuda, entre em contato com o DEV.")
        await ctx.reply(embed=embedVar)
        return

    if n == 0:
        embedVar = discord.Embed(title="⚠️Error 404", description="Não foi encontrado o seu registro em nosso servidor!", colour=discord.Colour.random())
        embedVar.set_author(name=ctx.author.name, icon_url=ctx.author.avatar_url)
        embedVar.set_footer(text="Por favor, entre em contato com o DEV.")
        await ctx.send(embed = embedVar)
        return

    for row in wsr.iter_rows(wsr.min_row, wsr.max_row):
        for cell in row:
            if cell.value == member_id:
                ws[f'{get_column_letter(cell.column + 8)}{(cell.row)}'] = "Não participante"
                wb.save(file)

    for role in ctx.guild.roles:
        if role.name == f"{clan_}":
            await ctx.author.remove_roles(role)

    embedVar = discord.Embed(title="✔️ Você saiu do clan com sucesso!", description = f"Agora você nao faz mais parte do clan {clan_}")
    embedVar.set_author(name=ctx.author.name, icon_url=ctx.author.avatar_url)
    embedVar.set_footer(text="➝ Se precisar de ajuda, entre em contato com o DEV.")
    await ctx.message.reply(embed = embedVar)

    return

@bot.command(name="clan_expulsar")
async def clan_expulsar(ctx, member:discord.Member):

    if str(datetime.datetime.now().hour) == '8':
        warning = await ctx.reply("Estamos atualizando o nosso banco de dados. Por favor, tente novamente mais tarde\n_Esse processo dura das 05:00 até às 06:00_\nCaso você ache que isso seja um erro, entre em contato com o DEV.")
        time.sleep(10)
        await warning.delete()
        await ctx.message.delete()
        return

    id = ctx.author.id
    member_id = "`" + str(id) + "`"

    file = "banco de dados.xlsx"
    wb = openpyxl.load_workbook(filename=file)
    ws = wb.worksheets[0]
    wsr = wb.active
    n = 0

    for row in wsr.iter_rows(wsr.min_row, wsr.max_row):
        for cell in row:
            if cell.value == member_id:
                clan = ws.cell(row=cell.row, column=9).value
                n += 1
            else:
                n += 0

    if clan == "Não participante":
        embedVar = discord.Embed(title="⚠️Error 401", description="Você não participa de um clan", colour=discord.Colour.random())
        embedVar.set_author(name=ctx.author.name, icon_url=ctx.author.avatar_url)
        embedVar.set_footer(text="➝ Se precisar de ajuda, entre em contato com o DEV.")
        await ctx.reply(embed = embedVar)
        return
    
    if not "Líder" in clan:
        embedVar = discord.Embed(title="⚠️Error 403", description="Você não é líder do seu clan", colour=discord.Colour.random())
        embedVar.set_author(name=ctx.author.name, icon_url=ctx.author.avatar_url)
        embedVar.set_footer(text="➝ Se precisar de ajuda, entre em contato com o DEV.")
        await ctx.reply(embed=embedVar)
        return

    if n == 0:
        embedVar = discord.Embed(title="⚠️Error 404", description="Não foi encontrado o seu registro em nosso servidor!", colour=discord.Colour.random())
        embedVar.set_author(name=ctx.author.name, icon_url=ctx.author.avatar_url)
        embedVar.set_footer(text="Por favor, entre em contato com o DEV.")
        await ctx.send(embed = embedVar)
        return
    
    clan_ = clan[:-8]

    member_id = "`" + str(member.id) + "`"

    n = 0

    for row in wsr.iter_rows(wsr.min_row, wsr.max_row):
        for cell in row:
            if cell.value == member_id:
                clan = ws.cell(row=cell.row, column=9).value
                n += 1
            else:
                n += 0

    if clan == "Não participante":
        embedVar = discord.Embed(title="⚠️Error 401", description="O usuário em questão não participa de nenhum clan", colour=discord.Colour.random())
        embedVar.set_author(name=ctx.author.name, icon_url=ctx.author.avatar_url)
        embedVar.set_footer(text="➝ Se precisar de ajuda, entre em contato com o DEV.")
        await ctx.reply(embed = embedVar)
        return
    
    if clan != clan_:
        embedVar = discord.Embed(title="⚠️Error 401", description="O usuário em questão não participa do seu clan", colour=discord.Colour.random())
        embedVar.set_author(name=ctx.author.name, icon_url=ctx.author.avatar_url)
        embedVar.set_footer(text="➝ Se precisar de ajuda, entre em contato com o DEV.")
        await ctx.reply(embed = embedVar)
        return
    
    if n == 0:
        embedVar = discord.Embed(title="⚠️Error 404", description="Não foi encontrado o registro do usuário em questão em nosso servidor!", colour=discord.Colour.random())
        embedVar.set_author(name=ctx.author.name, icon_url=ctx.author.avatar_url)
        embedVar.set_footer(text="Por favor, entre em contato com o DEV.")
        await ctx.send(embed = embedVar)
        return

    for row in wsr.iter_rows(wsr.min_row, wsr.max_row):
        for cell in row:
            if cell.value == member_id:
                ws[f'{get_column_letter(cell.column + 8)}{(cell.row)}'] = "Não participante"
                wb.save(file)
    
    for role in ctx.guild.roles:
        if role.name == f"{clan_}":
            await member.remove_roles(role)
    
    embedVar = discord.Embed(title=f"✔️ Você expulsou o {member.name} com sucesso", colour = discord.Colour.random())
    embedVar.set_author(name=ctx.author.name, icon_url=ctx.author.avatar_url)
    embedVar.set_footer(text="Por favor, entre em contato com o DEV.")
    await ctx.send(embed = embedVar)
    return

@bot.command(name="ajuda")
async def ajuda(ctx):
    check_1 = discord.utils.get(ctx.guild.roles, id=947242380040478761)
    if check_1 in ctx.author.roles:
        embedVar = discord.Embed(title="🤖 Ajuda do Bot State of War", description="Uma lista de ajuda feita para administradores do State of War\nSelecione o tipo de ajuda que você deseja receber nas opções abaixo!", colour=discord.Colour.random())
        embedVar.add_field(name="`!ajuda`", value="Mostra uma tabela de como usar os comandos disponiveis do bot.")
        embedVar.add_field(name="`!code [CODIGO]`", value="Vincula a sua Steam ao nosso sistema. Isso nos ajudará a otimizar o nosso suporte dinâmico. Recomendamos a todos vincularem suas contas.")
        embedVar.add_field(name="`!perfil`", value="Mostra como está a sua ficha em nosso servidor")
        embedVar.set_author(name=ctx.author.name, icon_url=ctx.author.avatar_url)
        embedVar.set_footer(text="➝ Se precisar de ajuda, entre em contato com o DEV.")
        mensagem = await ctx.reply(embed=embedVar, components=[Select(placeholder="Ajudas", custom_id="help", options=[
            SelectOption(label="Administração", value="adm", emoji="⚙️"),
            SelectOption(label="Kill/Death Ratio", value="kdr", emoji="⚜️"),
            SelectOption(label="Clan", value="clan", emoji="🛡️"),
            SelectOption(label="Música", value="musica", emoji="🎵"),
            SelectOption(label="Suspeito", value="suspeito", emoji="🕵️‍♂️")
        ])])

        interaction = await bot.wait_for('select_option', check=lambda inter: inter.custom_id == 'help' and inter.user == ctx.author)
        if interaction.values[0] == "adm":
            embedVar = discord.Embed(title="⚙️ Administração", description="Uma lista com comandos administrativos.", colour=discord.Colour.random())
            embedVar.set_author(name=ctx.author.name, icon_url=ctx.author.avatar_url)
            embedVar.add_field(name="`!ficha [@usuario]`", value="Mostra a ficha administrativa do usuário em questão.")
            embedVar.add_field(name="`!vipadd [@usuario]`", value="Adiciona uma vantagem para o usuário em questão.")
            embedVar.add_field(name="`!vipmod [@usuario]`", value="Modifica alguma vantagem do usuário em questão.")
            embedVar.add_field(name="`!punir [@usuario]`", value="Adiciona uma punição para o usuário em questão (infração, advertência ou banimento)")
            embedVar.add_field(name="`!enquete [TITULO]/[CONTEUDO]`", value="Cria uma enquete dinâmica. Os parametros Titulo/Conteudo são obrigatórios. **O Título e o Conteudo devem ser separados por '/'.**")
            embedVar.add_field(name="`!registrarall`", value="Registra todos os usuários novamente no banco de dados **(consulte o DEV antes de usar esse comando, pode gerar problemas em todo o banco de dados)**.")
            embedVar.add_field(name="`!da`", value="Anuncia a mensagem de doação do servidor. Usar quando o botão 'Saiba mais' apresentar alguma falha.")
            embedVar.add_field(name="`!regras_1`", value="Anuncia as regras de Chernarus que estão na memória do bot.", inline=True)
            embedVar.add_field(name="`!regras_2`", value="Anuncia as regras de Namalsk que estão na memória do bot.", inline=True)
            embedVar.add_field(name="`!giverole [CARGO]`", value="Da um [CARGO] em especifico para todos os membros do servidor.", inline=True)
            embedVar.add_field(name="`!removerole [CARGO]`", value="Remove um [CARGO] em especifico de todos os membros do servidor.", inline=True)
            embedVar.add_field(name="`!pegueseucargo`", value="Anuncia a mensagem para pegar cargos de Chernarus e/ou Namalsk no canal em questão.", inline=True)
            embedVar.add_field(name="`!valido`", value="Valida um raid (usado apenas em chats destinados a solicitação de prova de raid. Deleta o chat após 24 horas.)", inline=True)
            embedVar.add_field(name="`!invalido`", value="Invalida um raid (usado apenas em chats destinados a solicitação de prova de raid. Deleta o chat após 24 horas.)", inline=True)
            embedVar.add_field(name="`!purge`", value="Apaga TODAS as mensagens do canal de texto ao qual o comando foi enviado. Não existe forma de desfazer esse comando.", inline=True)
            embedVar.set_footer(text="➝ Se precisar de ajuda, entre em contato com o DEV.")
            await mensagem.edit(embed=embedVar)
        elif interaction.values[0] == "clan":
            embedVar = discord.Embed(title="🛡️ Clan", description="Uma lista com comandos para clan.", colour=discord.Colour.random())
            embedVar.set_author(name=ctx.author.name, icon_url=ctx.author.avatar_url)
            embedVar.add_field(name="`!clan_criar [TAG]`", value="Cria um clã com a sua TAG. Os parametros R,G,B podem são opcionais e, caso deixados em branco, a cor do seu clan será alatória.", inline=False)
            embedVar.add_field(name="`!clan_excluir`", value="Exclui o clan ao qual você possuí cargo de líder.", inline=False)
            embedVar.add_field(name="`!clan_convidar [@usuario]`", value="Envia um covite no privado do usuário em questão ao qual você deseja convidar para o seu clan (um usuario por vez).\nÉ importante que esse usuário em questão permita, nas configurações do Discord, receber mensagens de bots.",inline=False)
            embedVar.add_field(name="`!clan_sair`", value="Sai do clan em que você está participando. Líderes não podem sair do próprio clan.", inline=False)
            embedVar.add_field(name="`!clan_expulsar [@usuario]`", value="Expulsa o usuário em questão. Comando disponível apenas para os líderes de clan.", inline=False)
            embedVar.set_footer(text="➝ Se precisar de ajuda, entre em contato com o DEV.")
            await mensagem.edit(embed=embedVar)
        elif interaction.values[0] == "kdr":
            embedVar = discord.Embed(title="⚜️ Kill/Death ratio", description="Uma lista de comandos para analisar o seu KDr.", colour = discord.Colour.random())
            embedVar.set_author(name=ctx.author.name, icon_url=ctx.author.avatar_url)
            embedVar.add_field(name="`!kdr`", value="Mostra, publicamente, o seu KDr do Chernarus/Namalsk.", inline=False)
            embedVar.add_field(name="`!kdr_pv`", value="Envia, no seu privado, o seu KDr do Chernarus/Namalsk.", inline=False)
            embedVar.add_field(name="`!top_kills`", value="Mostra os 10 melhores tiros de jogadores.", inline=False)
            await mensagem.edit(embed=embedVar)
        elif interaction.values[0] == "musica":
            embedVar = discord.Embed(title="🎵 Música", description="Uma lista de comandos para gerenciar músicas.", colour = discord.Colour.random())
            embedVar.set_author(name=ctx.author.name, icon_url=ctx.author.avatar_url)
            embedVar.add_field(name="`!entrar`", value="Faz o Bot conectar-se a sua atual call de voz.", inline=False)
            embedVar.add_field(name="`!sair`", value="Faz o Bot sair da chamada ao qual ele está conectado.", inline=False)
            embedVar.add_field(name="`!agora`", value="Mostra a música que está tocando atualmente no Bot", inline=False)
            embedVar.add_field(name="`!skip`", value="Pula a música que está atualmente tocando no Bot.", inline=False)
            embedVar.add_field(name="`!fila`", value="Mostra a fila das músicas ainda a serem tocadas.", inline=False)
            embedVar.add_field(name="`!loop`", value="Coloca a música ou a playlist em loop/Tira a música ou a playlist de loop.", inline=False)
            embedVar.add_field(name="`!remover [NÚMERO]`", value="Remove a música da fila.", inline=False)
            embedVar.add_field(name="`!play [NOME OU LINK]`", value="Adiciona uma música na fila de reprodução.", inline=False)
            await mensagem.edit(embed=embedVar)
        elif interaction.values[0] == "suspeito":
            embedVar = discord.Embed(title="🕵️‍♂️ Suspeito", description="Uma lista de comandos para gerenciar os suspeitos.", colour = discord.Colour.random())
            embedVar.set_author(name=ctx.author.name, icon_url=ctx.author.avatar_url)
            embedVar.add_field(name="`!suspeito_adicionar [NICK]/[STEAM_ID]/[MOTIVO]`", value="Adiciona um SteamID, Nick e um motivo no banco de dados.", inline=False)
            embedVar.add_field(name="`!suspeito_deletar [STEAM_ID]`", value="Deleta um SteamID do banco de dados.", inline=False)
            embedVar.add_field(name="`!suspeito_editar [STEAM_ID]`", value="Edita alguma inforamação de algum SteamID em questão", inline=False)
            embedVar.add_field(name="`!suspeito_lista`", value="Mostra toda a lista de suspeitos que ja foram adicionados no banco de dados.", inline=False)
            embedVar.add_field(name="`!suspeito_ficha`", value="Mostra a ficha de algum Steam ID em questão.", inline=False)
            await mensagem.edit(embed=embedVar)
    else:
        embedVar = discord.Embed(title="🤖 Ajuda do Bot State of War", description="Uma lista de ajuda feita para administradores do State of War\nSelecione o tipo de ajuda que você deseja receber nas opções abaixo!", colour=discord.Colour.random())
        embedVar.add_field(name="`!ajuda`", value="Mostra uma tabela de como usar os comandos disponiveis do bot.")
        embedVar.add_field(name="`!code [CODIGO]`", value="Vincula a sua Steam ao nosso sistema. Isso nos ajudará a otimizar o nosso suporte dinâmico. Recomendamos a todos vincularem suas contas.")
        embedVar.set_author(name=ctx.author.name, icon_url=ctx.author.avatar_url)
        embedVar.set_footer(text="➝ Se precisar de ajuda, entre em contato com o DEV.")
        mensagem = await ctx.reply(embed=embedVar, components=[Select(placeholder="Ajudas", custom_id="help", options=[
            SelectOption(label="Clan", value="clan", emoji="🛡️"),
            SelectOption(label="Kill/Death Ratio", value="kdr", emoji="⚜️"),
            SelectOption(label="Música", value="musica", emoji="🎵")
        ])])
        interaction = await bot.wait_for('select_option', check=lambda inter: inter.custom_id == 'help' and inter.user == ctx.author)
        if interaction.values[0] == "clan":
            embedVar = discord.Embed(title="🛡️ Clan", description="Uma lista com comandos para clan.", colour=discord.Colour.random())
            embedVar.set_author(name=ctx.author.name, icon_url=ctx.author.avatar_url)
            embedVar.add_field(name="`!clan_criar [TAG]`", value="Cria um clã com a sua TAG. Os parametros R,G,B podem são opcionais e, caso deixados em branco, a cor do seu clan será alatória.", inline=False)
            embedVar.add_field(name="`!clan_excluir`", value="Exclui o clan ao qual você possuí cargo de líder.", inline=False)
            embedVar.add_field(name="`!clan_convidar [@usuario]`", value="Envia um covite no privado do usuário em questão ao qual você deseja convidar para o seu clan (um usuário por vez).\nÉ importante que esse usuário em questão permita, nas configurações do Discord, receber mensagens de bots.", inline=False)
            embedVar.add_field(name="`!clan_sair`", value="Sai do clan em que você está participando. Líderes não podem sair do próprio clan.", inline=False)
            embedVar.add_field(name="`!clan_expulsar [@usuario]`", value="Expulsa o usuário em questão. Comando disponível apenas para os líderes de clan.", inline=False)
            embedVar.set_footer(text="➝ Se precisar de ajuda, entre em contato com o DEV.")
            await mensagem.edit(embed=embedVar)
        elif interaction.values[0] == "musica":
            embedVar = discord.Embed(title="🎵 Música", description="Uma lista de comandos para gerenciar músicas.", colour = discord.Colour.random())
            embedVar.set_author(name=ctx.author.name, icon_url=ctx.author.avatar_url)
            embedVar.add_field(name="`!entrar`", value="Faz o Bot conectar-se a sua atual call de voz.", inline=False)
            embedVar.add_field(name="`!sair`", value="Faz o Bot sair da chamada ao qual ele está conectado.", inline=False)
            embedVar.add_field(name="`!agora`", value="Mostra a música que está tocando atualmente no Bot", inline=False)
            embedVar.add_field(name="`!skip`", value="Pula a música que está atualmente tocando no Bot.", inline=False)
            embedVar.add_field(name="`!fila`", value="Mostra a fila das músicas ainda a serem tocadas.", inline=False)
            embedVar.add_field(name="`!loop`", value="Coloca a música ou a playlist em loop/Tira a música ou a playlist de loop.", inline=False)
            embedVar.add_field(name="`!remover [NÚMERO]`", value="Remove a música da fila.", inline=False)
            embedVar.add_field(name="`!play [NOME OU LINK]`", value="Adiciona uma música na fila de reprodução.", inline=False)
            await mensagem.edit(embed=embedVar)
        elif interaction.values[0] == "kdr":
            embedVar = discord.Embed(title="⚜️ Kill/Death ratio", description="Uma lista de comandos para analisar o seu KDr.", colour = discord.Colour.random())
            embedVar.set_author(name=ctx.author.name, icon_url=ctx.author.avatar_url)
            embedVar.add_field(name="`!kdr`", value="Mostra, publicamente, o seu KDr do Chernarus/Namalsk.", inline=False)
            embedVar.add_field(name="`!kdr_pv`", value="Envia, no seu privado, o seu KDr do Chernarus/Namalsk.", inline=False)
            embedVar.add_field(name="`!top_kills`", value="Mostra os 10 melhores tiros de jogadores.", inline=False)
            await mensagem.edit(embed=embedVar)
    return

@bot.command(name="registrar")
async def registrar(ctx, member:discord.Member):

    if str(datetime.datetime.now().hour) == '8':
        warning = await ctx.reply("Estamos atualizando o nosso banco de dados. Por favor, tente novamente mais tarde\n_Esse processo dura das 05:00 até às 06:00_\nCaso você ache que isso seja um erro, entre em contato com o DEV.")
        time.sleep(10)
        await warning.delete()
        await ctx.message.delete()
        return

    check_1 = discord.utils.get(ctx.guild.roles, id=947242380040478761) #ID DE ADMINISTRADOR SUPREMO

    if check_1 in ctx.author.roles:
        pass
    else:
        warning = await ctx.reply("Você não tem permissão para usar isso!")
        time.sleep(8)
        await warning.delete()
        await ctx.message.delete()
        return

    file = "banco de dados.xlsx"
    wb = openpyxl.load_workbook(filename=file)
    ws = wb.worksheets[0]
    wsr = wb.active

    id = member.id
    id = str(id)
    id = "`"+id+"`"
    name = member.name
    discriminator = member.discriminator
    nickname = name + "#" + discriminator
    ws[f'{get_column_letter(ws.min_column + 0)}{(ws.max_row + 1)}'] = id
    ws[f'{get_column_letter(ws.min_column + 1)}{(ws.max_row)}'] = nickname
    ws[f'{get_column_letter(ws.min_column + 5)}{(ws.max_row)}'] = "0" #INFRAÇÕES
    ws[f'{get_column_letter(ws.min_column + 6)}{(ws.max_row)}'] = "0" #ADVERTÊNCIAS
    ws[f'{get_column_letter(ws.min_column + 7)}{(ws.max_row)}'] = "0" #BANIMENTOS
    ws[f'{get_column_letter(ws.min_column + 8)}{(ws.max_row)}'] = "Não participante" #CLÃ
    ws[f'{get_column_letter(ws.min_column + 9)}{(ws.max_row)}'] = "❌" #SEGURO DE VEICULOS
    ws[f'{get_column_letter(ws.min_column + 10)}{(ws.max_row)}'] = "0" #QUANTOS DIAS RESTANTES
    ws[f'{get_column_letter(ws.min_column + 11)}{(ws.max_row)}'] = "Indefinido" #QUANTOS SEGUROS RESTANTES DE VEICULOS
    ws[f'{get_column_letter(ws.min_column + 12)}{(ws.max_row)}'] = "❌" #SEGURO DE HELICOPTEROS
    ws[f'{get_column_letter(ws.min_column + 13)}{(ws.max_row)}'] = "0" #QUANTOS DIAS RESTANTES
    ws[f'{get_column_letter(ws.min_column + 14)}{(ws.max_row)}'] = "Indefinido" #QUANTOS SEGUROS RESTANTES DE HELICOPTEROS
    ws[f'{get_column_letter(ws.min_column + 15)}{(ws.max_row)}'] = "❌" #POSSUI ALGUMA CONSTRUÇÃO
    ws[f'{get_column_letter(ws.min_column + 16)}{(ws.max_row)}'] = "0" #QUANTOS DIAS RESTANTES DA CONSTRUÇÃO
    ws[f'{get_column_letter(ws.min_column + 17)}{(ws.max_row)}'] = "❌" #POSSUI FILA PRIORITARIA
    ws[f'{get_column_letter(ws.min_column + 18)}{(ws.max_row)}'] = "0" #QUANTOS DIAS RESTANTES 
    wb.save(file)

@bot.command(name="daypass")
async def daypass(ctx):

    if str(datetime.datetime.now().hour) == '5':
        warning = await ctx.reply("Estamos atualizando o nosso banco de dados. Por favor, tente novamente mais tarde\n_Esse processo dura das 05:00 até às 06:00_\nCaso você ache que isso seja um erro, entre em contato com o DEV.")
        time.sleep(10)
        await warning.delete()
        await ctx.message.delete()
        return

    check_1 = discord.utils.get(ctx.guild.roles, id=947242380040478761) #ID DE ADMINISTRADOR SUPREMO

    if check_1 in ctx.author.roles:
        pass
    else:
        warning = await ctx.reply("Você não tem permissão para usar isso!")
        time.sleep(8)
        await warning.delete()
        await ctx.message.delete()
        return

    file = "banco de dados.xlsx"
    wb = openpyxl.load_workbook(filename=file)
    ws = wb.worksheets[0]
    wsr = wb.active

    async def update(coluna):
        if coluna == 11:
            beneficio = "Seguro de Veiculos"
        if coluna == 14:
            beneficio = "Seguro de Helicopteros"
        if coluna == 17:
            beneficio = "Base VIP"
        if coluna == 19:
            beneficio = "Fila prioritária"

        status = "normal"

        for row in wsr.iter_cols(min_col=coluna, max_col=coluna, min_row=1):
                for cell in row:
                    if cell.value != '0' and cell.value != None:
                        idd = ws.cell(row=cell.row, column=1).value
                        idd = idd.replace("`","")
                        antigo = int(cell.value)
                        novo = antigo - 1
                        ws[f'{get_column_letter(cell.column)}{(cell.row)}'] = str(novo) #TEMPO RESTANTE DA VILA PRIORITARIA
                        if str(novo) == '0':
                            ws[f'{get_column_letter(cell.column -1)}{(cell.row)}'] = '❌'
                            status = "acabou"
                        wb.save(file)
                        if str(novo) == '3':
                            status = "quase"

        if status == "quase":
            user = await bot.fetch_user(idd)
            embedVar = discord.Embed(title="⚠️ Atenção!", description=f"O benefício de `{beneficio}` de {user.name}, portador do Discord ID {idd} está a 3 dias do seu fim", colour=discord.Colour.random())
            embedVar.set_footer(text="Se precisar de ajuda, entre em contato com o DEV.")
            await channel.send(embed = embedVar)

            embedVar = discord.Embed(title="⚠️ Atenção!", description=f"Faltam 3 dias para seu benefício de `{beneficio}` acabar! Você pode renovar criando um ticket na aba de Financeiro!", colour=discord.Colour.random())
            embedVar.set_footer(text="Se precisar de ajuda, entre em contato com o DEV.")
            await user.send(embed = embedVar)
        
        if status == "acabou":
            user = await bot.fetch_user(idd)
            embedVar = discord.Embed(title="⚠️ Atenção!", description=f"O benefício de `{beneficio}` de {user.name}, portador do Discord ID {idd} **acabou**!", colour=discord.Colour.random())
            embedVar.set_footer(text="Se precisar de ajuda, entre em contato com o DEV.")
            await channel.send(embed = embedVar)

            embedVar = discord.Embed(title="⚠️ Atenção!", description=f"O seu beneficio de `{beneficio}` acabou! Você pode renovar criando um ticket na aba de Financeiro!", colour=discord.Colour.random())
            embedVar.set_footer(text="Se precisar de ajuda, entre em contato com o DEV.")
            await user.send(embed = embedVar)
        
    await update(11)
    await update(14)
    await update(17)
    await update(19)

    return

@bot.command(name="giverole")
async def giverole(ctx, role:discord.Role):

    check_1 = discord.utils.get(ctx.guild.roles, id=947242380040478761) #ID DE ADMINISTRADOR SUPREMO

    if check_1 in ctx.author.roles:
        pass
    else:
        warning = await ctx.reply("Você não tem permissão para usar isso!")
        time.sleep(8)
        await warning.delete()
        await ctx.message.delete()
        return


    for member in ctx.guild.members:
        await member.add_roles(role)
    
    await ctx.send("Feito!")

    return  

@bot.command(name="removerole")
async def removerole(ctx, role:discord.Role):

    check_1 = discord.utils.get(ctx.guild.roles, id=947242380040478761) #ID DE ADMINISTRADOR SUPREMO

    if check_1 in ctx.author.roles:
        pass
    else:
        warning = await ctx.reply("Você não tem permissão para usar isso!")
        time.sleep(8)
        await warning.delete()
        await ctx.message.delete()
        return


    for member in ctx.guild:
        await member.remove_roles(role)
    return    

@bot.command(name="pegueseucargo")
async def pegueseucargo(ctx):
    check_1 = discord.utils.get(ctx.guild.roles, id=947242380040478761) #ID DE ADMINISTRADOR SUPREMO

    if check_1 in ctx.author.roles:
        pass
    else:
        warning = await ctx.reply("Você não tem permissão para usar isso!")
        time.sleep(8)
        await warning.delete()
        await ctx.message.delete()
        return
    
    embedVar = discord.Embed(title="✅ Pegue seu cargo aqui!", description="Para controlar melhor o servidor e torna-lo mais dinâmico, pegue os cargos dos servidores que você joga! Assim, você receberá acesso aos canais relativos a sua necessidade.", colour=discord.Colour.random())
    embedVar.add_field(name="Cargos:", value="☢️ Chernarus\n❄️ Namalsk")
    embedVar.set_footer(text="➝ Se precisar de ajuda, entre em contato com o DEV.")
    message = await ctx.send(embed=embedVar)
    await ctx.message.delete()
    await message.add_reaction('☢️')
    await message.add_reaction('❄️')

@bot.command(name="raid_msg_chernarus")
async def raid_msg_chernarus(ctx):

    check_1 = discord.utils.get(ctx.guild.roles, id=947242380040478761) #ID DE ADMINISTRADOR SUPREMO

    if check_1 in ctx.author.roles:
        pass
    else:
        warning = await ctx.reply("Você não tem permissão para usar isso!")
        time.sleep(8)
        await warning.delete()
        await ctx.message.delete()
        return

    embedVar = discord.Embed(title="💥 Solicitação de prova de Raid", description="Para abrir uma solicitação de prova de Raid, clique no botão abaixo.", colour=discord.Colour.red())
    embedVar.add_field(name="Observações:", value="**・Somente crie solicitações se necessário.**\n・Você só poderá criar uma nova solicitação.\n・Espere a avaliação do Administrador para o seu caso.\n・Para agilizar o processo, mande a localização da base logo após criar a solicitação, deixando claro onde a base raidada se localiza.")
    embedVar.set_footer(text="➝ Se precisar de ajuda, entre em contato com o DEV")
    await ctx.send(embed=embedVar, components=[Button(label="Exigir prova de Raid", custom_id='raid_request_chernarus', style=ButtonStyle.red)])
    await ctx.message.delete()

@bot.command(name="raid_msg_namalsk")
async def raid_msg_namalsk(ctx):

    check_1 = discord.utils.get(ctx.guild.roles, id=947242380040478761) #ID DE ADMINISTRADOR SUPREMO

    if check_1 in ctx.author.roles:
        pass
    else:
        warning = await ctx.reply("Você não tem permissão para usar isso!")
        time.sleep(8)
        await warning.delete()
        await ctx.message.delete()
        return

    embedVar = discord.Embed(title="💥 Solicitação de prova de Raid", description="Para abrir uma solicitação de prova de Raid, clique no botão abaixo.", colour=discord.Colour.red())
    embedVar.add_field(name="Observações:", value="**・Somente crie solicitações se necessário.**\n・Você só poderá criar uma nova solicitação.\n・Espere a avaliação do Administrador para o seu caso.\n・Para agilizar o processo, mande a localização da base logo após criar a solicitação, deixando claro onde a base raidada se localiza.")
    embedVar.set_footer(text="➝ Se precisar de ajuda, entre em contato com o DEV")
    await ctx.send(embed=embedVar, components=[Button(label="Exigir prova de Raid", custom_id='raid_request_namalsk', style=ButtonStyle.red)])
    await ctx.message.delete()

@bot.command(name="valido")
async def valido(ctx):

    check_1 = discord.utils.get(ctx.guild.roles, id=947242380040478761) #ID DE ADMINISTRADOR SUPREMO

    if check_1 in ctx.author.roles:
        pass
    else:
        warning = await ctx.reply("Você não tem permissão para usar isso!")
        time.sleep(8)
        await warning.delete()
        await ctx.message.delete()
        return

    global id_channel_chernarus
    global id_requests_chernarus
    global id_channel_namalsk
    global id_requests_namalsk

    channel_id_ = ctx.channel.id

    if not channel_id_ in id_channel_chernarus:
        if not channel_id_ in id_channel_namalsk:
            warning = await ctx.send('Esse comando é exclusivo apenas para validar raids. Por favor, use esse comando em um chat de texto relacionado a aprovação de Raid')
            time.sleep(8)
            await ctx.message.delete()
            await warning.delete()
            return
    
    embedVar = discord.Embed(title="✅ Raid validado!", description="O raid foi analisado e aprovado pela administração do State of War e está de acordo com as normas de raid.\n_Esse canal poderá ser apagado a qualquer momento._", colour=discord.Colour.green())
    embedVar.set_footer(text="➝ Se precisar de ajuda, entre em contato com o DEV.")
    await ctx.send(embed=embedVar)

    try:
        id_channel_chernarus.remove(ctx.channel.id)
    except ValueError:
        pass

    try:
        id_channel_namalsk.remove(ctx.channel.id)
    except ValueError:
        pass

    await ctx.message.delete()

    requester = await ctx.channel.history(oldest_first=True, limit=1).flatten()
    requester = str(requester)
    requester = requester[13:]
    requester = requester.split(" ")
    requester = requester[0]
    requester = await ctx.channel.fetch_message(int(requester))
    requester = requester.content
    requester = requester[4:]
    requester = str(requester)
    requester = requester.split(" ")
    map = requester[1]
    requester = requester[0]

    try:
        id_requests_chernarus.remove(int(requester))
    except ValueError:
        pass
    try:
        id_requests_namalsk.remove(int(requester))
    except ValueError:
        pass

    if map == 'Chernarus':
        await ctx.channel.set_permissions(ctx.guild.get_role(961331532012863498), send_messages=False, read_messages=True, read_message_history=True)
    elif map == 'Namalsk':
        await ctx.channel.set_permissions(ctx.guild.get_role(961331570717917304), send_messages=False, read_messages=True, read_message_history=True)

@bot.command(name="invalido")
async def invalido(ctx):
    check_1 = discord.utils.get(ctx.guild.roles, id=947242380040478761) #ID DE ADMINISTRADOR SUPREMO

    if check_1 in ctx.author.roles:
        pass
    else:
        warning = await ctx.reply("Você não tem permissão para usar isso!")
        time.sleep(8)
        await warning.delete()
        await ctx.message.delete()
        return

    global id_channel_chernarus
    global id_requests_chernarus
    global id_channel_namalsk
    global id_requests_namalsk

    channel_id_ = ctx.channel.id

    if not channel_id_ in id_channel_chernarus:
        if not channel_id_ in id_channel_namalsk:
            warning = await ctx.send('Esse comando é exclusivo apenas para validar raids. Por favor, use esse comando em um chat de texto relacionado a aprovação de Raid')
            time.sleep(8)
            await ctx.message.delete()
            await warning.delete()
            return

    embedVar = discord.Embed(title="❌ Raid invalidado!", description="O raid foi analisado e desaprovado pela administração do State of War e não está de acordo com as normas de raid.\n_Esse canal poderá ser apagado a qualquer momento._\nA partir de agora, os administradores irão decidir o que poderá ser feito diante dessa situação problemática.", colour=discord.Colour.red())
    embedVar.set_footer(text="➝ Se precisar de ajuda, entre em contato com o DEV.")
    await ctx.send(embed=embedVar)

    try:
        id_channel_chernarus.remove(ctx.channel.id)
    except ValueError:
        pass

    try:
        id_channel_namalsk.remove(ctx.channel.id)
    except ValueError:
        pass

    await ctx.message.delete()

    requester = await ctx.channel.history(oldest_first=True, limit=1).flatten()
    requester = str(requester)
    requester = requester[13:]
    requester = requester.split(" ")
    requester = requester[0]
    requester = await ctx.channel.fetch_message(int(requester))
    requester = requester.content
    requester = requester[4:]
    requester = str(requester)
    requester = requester.split(" ")
    map = requester[1]
    requester = requester[0]

    try:
        id_requests_chernarus.remove(int(requester))
    except ValueError:
        pass
    try:
        id_requests_namalsk.remove(int(requester))
    except ValueError:
        pass

    if map == 'Chernarus':
        await ctx.channel.set_permissions(ctx.guild.get_role(961331532012863498), send_messages=False, read_messages=True, read_message_history=True)
    elif map == 'Namalsk':
        await ctx.channel.set_permissions(ctx.guild.get_role(961331570717917304), send_messages=False, read_messages=True, read_message_history=True)

@bot.event
async def on_button_click(interaction):

    global id_requests_chernarus
    global id_channel_chernarus

    global id_requests_namalsk
    global id_channel_namalsk

    guild = bot.get_guild(947237264596041728)

    if interaction.custom_id == 'raid_request_chernarus':
        if not interaction.author.id in id_requests_chernarus:

            await interaction.send("Foi aberto uma solicitação de raid em seu nome. Por favor, envie, no canal onde você foi mencionado, os dados da sua base")

            category = discord.utils.get(guild.categories, id=947237264596041729)

            channel = await guild.create_text_channel('💥・prova-de-raid', category=category, position=1001)

            id_channel_chernarus.append(channel.id)
            id_requests_chernarus.append(interaction.author.id)

            await channel.send(content=f"ID: {interaction.author.id} Chernarus")
            embedVar = discord.Embed(title=f"💥 Solicitação de raid", colour = discord.Colour.random())
            embedVar.add_field(name="Observações", value="・Quem raidou tem um prazo de 48h para postar a prova de raid.")
            embedVar.set_footer(text="➝ Se precisar de ajuda, entre em contato com o DEV")
            await channel.send(embed=embedVar)

            await channel.send(f"Solicitação de prova de raid aberta por {interaction.author.mention}")

        elif interaction.author.id in id_requests_chernarus:
            
            await interaction.send("Já existe uma solicitação de raid em seu nome em andamento. Aguarde o fechamento da mesma para você conseguir abrir outra solicitação.")
   
    if interaction.custom_id == 'raid_request_namalsk':
        if not interaction.author.id in id_requests_namalsk:

            await interaction.send("Foi aberto uma solicitação de raid em seu nome. Por favor, envie, no canal onde você foi mencionado, os dados da sua base")

            category = discord.utils.get(guild.categories, id=961334974823424030)

            channel = await guild.create_text_channel('💥・prova-de-raid', category=category, position=1001)

            id_channel_namalsk.append(channel.id)
            id_requests_namalsk.append(interaction.author.id)

            await channel.send(content=f"ID: {interaction.author.id} Namalsk")
            embedVar = discord.Embed(title=f"💥 Solicitação de raid", colour = discord.Colour.random())
            embedVar.add_field(name="Observações", value="・Quem raidou tem um prazo de 48h para postar a prova de raid.")
            embedVar.set_footer(text="➝ Se precisar de ajuda, entre em contato com o DEV")
            await channel.send(embed=embedVar)

            await channel.send(f"Solicitação de prova de raid aberta por {interaction.author.mention}")

        elif interaction.author.id in id_requests_namalsk:
            
            await interaction.send("Já existe uma solicitação de raid em seu nome em andamento. Aguarde o fechamento da mesma para você conseguir abrir outra solicitação.")
       
    if interaction.custom_id == 'more_info_vip':

        inicio = discord.Embed(title="Obrigado!", description="Obrigado por considerar faze uma doação ao nosso servidor!\nSegue abaixo uma lista com todos as vantagens das doações!", colour= discord.Colour.random())
        inicio.add_field(name="Mas... Como eu posso doar?", value="Abra um ticket e entre em contato com um Administrador, enviado o tipo de doação e o comprovante de pagamento.")
        inicio.set_footer(text="©️ Todos os direitos reservados.")
 
        await interaction.send("As mensagens estarão sendo enviadas no seu privado!")
        await interaction.user.send(embed=inicio)
        await interaction.user.send(file=discord.File("images/divisoria.gif"))
        await interaction.user.send("```diff\n-DOAÇÃO FILA PRIORITÁRIA       (R$20,00 MENSAL)\n-DOAÇÃO SEGURO DE VEÍCULOS     (R$30,00 MENSAL)\n-DOAÇÃO SEGURO DE HELICOPTEROS (R$50,00 MENSAL)\n-DOAÇÃO CARLOCKPICK  (UNIDADE) (R$10,00)```")
        await interaction.user.send(file=discord.File("images/divisoria.gif"))
        await interaction.user.send("```diff\n-DOAÇÃO PACK DE ARMAS 1\n-5 ARMAS DE RUSH\n-VALOR: R$10,00```")
        await interaction.user.send("```diff\n-DOAÇÃO PACK DE ARMAS 2\n-5 SNIPERS .338\n-VALOR: R$10,00```")
        await interaction.user.send("```diff\n-DOAÇÃO PACK DE ARMAS 3\n-10 ARMAS DE RUSH\n-VALOR: R$20,00```")
        await interaction.user.send("```diff\n-DOAÇÃO PACK DE ARMAS 4\n-10 SNIPERS .338\n-VALOR: R$20,00```")
        await interaction.user.send("```diff\n-DOAÇÃO PACK DE ARMAS CABULOSO\n-20 ARMAS DA SUA ESCOLHA\n-VALOR: R$50,00```")
        await interaction.user.send(file=discord.File("images/divisoria.gif"))
        await interaction.user.send(file=discord.File("images/1milhao.png"))
        await interaction.user.send(file=discord.File("images/divisoria.gif"))
        await interaction.user.send(file=discord.File("images/2milhoes.png"))
        await interaction.user.send(file=discord.File("images/divisoria.gif"))
        await interaction.user.send(file=discord.File("images/4milhoes.png"))
        await interaction.user.send(file=discord.File("images/divisoria.gif"))
        await interaction.user.send(file=discord.File("images/BUNKER PEQUENO.png"))
        await interaction.user.send(file=discord.File("images/divisoria.gif"))
        await interaction.user.send(file=discord.File("images/BUNKER GRANDE.png"))
        await interaction.user.send(file=discord.File("images/divisoria.gif"))
        await interaction.user.send(file=discord.File("images/PREDIO.png"))

@bot.event
async def on_message(message):

    await bot.process_commands(message)
    
    if message.channel.id == 960055351833677875:
        interaction = await bot.fetch_user(id)

    if message.author.id == 912310660669521921:
        return

    sugestao_banidas = ['inclinado','pescoço torto','torto','inclinar','inclinação','inclinacao','perninha quebrada','enclinadinha','enclinada','enclinacao','enclinação','pescoço','pescosso','cabecinha','cabeçinha','cabessinha','cabesinha','deitar para o lado','deitar pro lado','deitar de lado']
    sugestao_banidas2 = ['stun', 'shok', 'choque', 'impulsos', 'impulssos', 'inpulso', 'inpulsso', 'knock', 'knok', 'knock', 'chok']
    if message.channel.id == 966561948806307912:

        for word in sugestao_banidas:
            if word in message.content:
                warning = await message.reply("Nós não iremos colocar inclinação no servidor!")
                await message.delete()
                time.sleep(10)
                await warning.delete()
                return
        
        for word in sugestao_banidas2:
            if word in message.content:
                warning = await message.reply("Nós não pretendemos mudar os impactos dos tiros no servidor!")
                await message.delete()
                time.sleep(10)
                await warning.delete
                return

        guild = await bot.fetch_guild(947237264596041728)
        sim = await guild.fetch_emoji(963299957555752970)
        nao = await guild.fetch_emoji(963300068578963516)

        await message.add_reaction(sim)
        await message.add_reaction(nao)

    if message.channel.id == 961335605487345725:

        for word in sugestao_banidas:
            if word in message.content:
                warning = await message.reply("Nos não iremos colocar inclinação no servidor!")
                await message.delete()
                time.sleep(10)
                await warning.delete()
                return

        guild = await bot.fetch_guild(947237264596041728)
        sim = await guild.fetch_emoji(963299957555752970)
        nao = await guild.fetch_emoji(963300068578963516)

        await message.add_reaction(sim)
        await message.add_reaction(nao)

    if message.channel.id == 964755459766620220:
        embeds = message.embeds
        for embed in embeds:
            if 'joined' in str(embed.to_dict()):
                content = str(embed.to_dict())
                content = content.split("}")
                content = content[1]
                content = content[24:]
                id = content[:17]

                id = "`" + str(id) + "`"

                file = "suspeitos.xlsx"
                wb = openpyxl.load_workbook(filename=file)
                ws = wb.worksheets[0]
                wsr = wb.active

                for row in wsr.iter_rows(wsr.min_row, wsr.max_row):
                    for cell in row:
                        if cell.value == id:
                            channel = await bot.fetch_channel(960022037794025522)

                            embedVar = discord.Embed(title="⚠️ ALERTA!!! ⚠️", description=f"O JOGADOR COM O STEAMID {id} ENTROU NO JOGO E ESTÁ NA LISTA DE SUSPEITOS!!!", colour = discord.Colour.red())
                            embedVar.set_footer(text=f"Para saber mais informações sobre o suspeito, use `!suspeito_ficha {id}`\nCaso tenha alguma dúvida, entre em contato com o DEV.")
                            await channel.send(embed = embedVar)

    if message.channel.id == 961302672609267783: #namalsk kds
        embeds = message.embeds
        for embed in embeds:
            msg = str(embed.to_dict())

            try:
                content = msg
                content = content.split("}")
                informacao = content[1]
                content = informacao.split('\\n')
                content = content[1]
                content = content.replace("[","")
                content = content.replace("]","")
                steam_id_death = str(content)
                steam_id_death = "`" + steam_id_death + "`"
            except Exception:
                pass
            
            try:
                content = msg
                content = content.split("}")
                informacao_2 = content[2]
                content = informacao_2.split('\\n')
                content = content[1]
                content = content.replace("[","")
                content = content.replace("]","")
                steam_id_killer = str(content)
                steam_id_killer = "`" + steam_id_killer + "`"
            except Exception:
                pass

            try:
                content = msg
                content = content.split("}")
                informacao_3 = content[3]
                content = informacao_3.split("\\nPosition:")
                content = content[2]
                content = content.split("from")
                content = content[1]
                content = content.split("meters")
                content = content[0]
                content = content.replace("[","")
                content = content.replace("]","")
                distancia = float(content)
                distancia = round(distancia, 3)
            except Exception:
                pass
            
            file = "kds_namalsk.xlsx"
            wb = openpyxl.load_workbook(filename=file)
            ws = wb.worksheets[0]
            wsr = wb.active

            for row in wsr.iter_rows(wsr.min_row, wsr.max_row):
                for cell in row:
                    if cell.value == steam_id_death:
                        mortes = ws.cell(row=cell.row, column = 3).value
                        new_mortes = int(mortes) + 1
                        ws[f'{get_column_letter(cell.column + 1)}{(cell.row)}'] = str(new_mortes)
                        wb.save(file)
                    try:
                        if cell.value == steam_id_killer:
                            kills = ws.cell(row=cell.row, column = 4).value
                            distancia_anterior = ws.cell(row=cell.row, column=5).value
                            new_kills = int(kills) + 1
                            ws[f'{get_column_letter(cell.column + 2)}{(cell.row)}'] = str(new_kills)
                            wb.save(file)
                            if float(distancia_anterior) < float(distancia):
                                ws[f'{get_column_letter(cell.column + 3)}{(cell.row)}'] = str(distancia)
                                wb.save(file)
                    except Exception:
                        pass

    if message.channel.id == 961302180340596826: #chernarus kds
        embeds = message.embeds
        for embed in embeds:

            msg = str(embed.to_dict())

            try:
                content = msg
                content = content.split("}")
                informacao = content[1]
                content = informacao.split('\\n')
                content = content[1]
                content = content.replace("[","")
                content = content.replace("]","")
                steam_id_death = str(content)
                steam_id_death = "`" + steam_id_death + "`"
            except Exception:
                pass

            try:
                content = msg
                content = content.split("}")
                informacao_2 = content[2]
                content = informacao_2.split('\\n')
                content = content[1]
                content = content.replace("[","")
                content = content.replace("]","")
                steam_id_killer = str(content)
                steam_id_killer = "`" + steam_id_killer + "`"
            except Exception:
                pass

            try:
                content = msg
                content = content.split("}")
                informacao_3 = content[3]
                content = informacao_3.split("\\nPosition:")
                content = content[2]
                content = content.split("from")
                content = content[1]
                content = content.split("meters")
                content = content[0]
                content = content.replace("[","")
                content = content.replace("]","")
                distancia = float(content)
                distancia = round(distancia, 3)
            except Exception:
                pass
            
            file = "kds_chernarus.xlsx"
            wb = openpyxl.load_workbook(filename=file)
            ws = wb.worksheets[0]
            wsr = wb.active

            for row in wsr.iter_rows(wsr.min_row, wsr.max_row):
                for cell in row:
                    if cell.value == steam_id_death:
                        mortes = ws.cell(row=cell.row, column = 3).value
                        new_mortes = int(mortes) + 1
                        ws[f'{get_column_letter(cell.column + 1)}{(cell.row)}'] = str(new_mortes)
                        wb.save(file)
                    try:
                        if cell.value == steam_id_killer:
                            kills = ws.cell(row=cell.row, column = 4).value
                            distancia_anterior = ws.cell(row=cell.row, column=5).value
                            new_kills = int(kills) + 1
                            ws[f'{get_column_letter(cell.column + 2)}{(cell.row)}'] = str(new_kills)
                            wb.save(file)
                            if float(distancia_anterior) < float(distancia):
                                ws[f'{get_column_letter(cell.column + 3)}{(cell.row)}'] = str(distancia)
                                wb.save(file)
                    except Exception:
                        pass

@bot.event
async def on_member_remove(member):
    guild = bot.get_guild(947237264596041728)
    member_count = guild.member_count
    activity = discord.Activity(name="Sobreviventes ➝ ({})".format(member_count), type=1)
    await bot.change_presence(status=discord.Status.online, activity=activity)

    channel = bot.get_channel(959881836241248276)

    author = member.name
    pic = member.avatar_url

    embedVar = discord.Embed(title=f"🧟‍♂️|**{author}**| foi mordido por um zumbi!", description=f"Infelizmente, {author} não está mais conosco. Agora somos {member_count} sobreviventes.", colour = discord.Colour.random())
    embedVar.set_author(name=author, icon_url=pic)
    embedVar.set_thumbnail(url=pic)
    embedVar.set_image(url="https://i.imgur.com/1td2WQv.png")
    embedVar.set_footer(text="©️ Todos os direitos reservados.")

    await channel.send(embed=embedVar)
    return

@bot.event
async def on_member_join(member):

    if str(datetime.datetime.now().hour) == '8':
        time.sleep(1000)

    file = "banco de dados.xlsx"
    wb = openpyxl.load_workbook(filename=file)
    ws = wb.worksheets[0]
    wsr = wb.active

    role = discord.utils.get(member.guild.roles, name='Sobreviventes')

    await member.add_roles(role)
    guild = bot.get_guild(947237264596041728)
    member_count = guild.member_count
    activity = discord.Activity(name="Sobreviventes ➝ ({})".format(member_count), type=1)
    await bot.change_presence(status=discord.Status.online, activity=activity)

    ID_REGRAS = 947238415013916692
    channel = bot.get_channel(947237264596041730) 
    guild = bot.get_guild(947237264596041728)
    member_count = guild.member_count
    author = member.name
    pic = member.avatar_url

    embedVar = discord.Embed(title="<a:welcome:958517359667187783> | Bem-vindo(a)! | <a:welcome:958517359667187783>", description="__**{}**__, seja bem-vindo ao State of War! Você é {}º membro a entrar na comunidade!".format((author),(member_count)), colour = discord.Colour.random())
    embedVar.set_author(name=author, icon_url=pic)
    embedVar.set_footer(text="©️ Todos os direitos reservados.")
    embedVar.set_image(url="https://i.imgur.com/pNBsQTi.jpg")
    embedVar.add_field(name="Evite punições <a:siren:963299760662544416> ", value="Leia as regras em <#{}> e fique por dentro das diretrizes da nossa comunidade.".format(ID_REGRAS), inline=False)
    embedVar.set_thumbnail(url=pic)
    await channel.send(embed=embedVar)

    id = member.id
    id = str(id)
    id = "`"+id+"`"
    name = member.name
    discriminator = member.discriminator
    nickname = name + "#" + discriminator

    ws[f'{get_column_letter(ws.min_column + 0)}{(ws.max_row + 1)}'] = id
    ws[f'{get_column_letter(ws.min_column + 1)}{(ws.max_row)}'] = nickname
    ws[f'{get_column_letter(ws.min_column + 5)}{(ws.max_row)}'] = "0" #INFRAÇÕES
    ws[f'{get_column_letter(ws.min_column + 6)}{(ws.max_row)}'] = "0" #ADVERTÊNCIAS
    ws[f'{get_column_letter(ws.min_column + 7)}{(ws.max_row)}'] = "0" #BANIMENTOS
    ws[f'{get_column_letter(ws.min_column + 8)}{(ws.max_row)}'] = "Não participante" #CLAN
    ws[f'{get_column_letter(ws.min_column + 9)}{(ws.max_row)}'] = "❌" #SEGURO DE VEICULOS
    ws[f'{get_column_letter(ws.min_column + 10)}{(ws.max_row)}'] = "0" #QUANTOS DIAS RESTANTES
    ws[f'{get_column_letter(ws.min_column + 11)}{(ws.max_row)}'] = "Indefinido" #QUANTOS SEGUROS RESTANTES DE VEICULOS
    ws[f'{get_column_letter(ws.min_column + 12)}{(ws.max_row)}'] = "❌" #SEGURO DE HELICOPTEROS
    ws[f'{get_column_letter(ws.min_column + 13)}{(ws.max_row)}'] = "0" #QUANTOS DIAS RESTANTES
    ws[f'{get_column_letter(ws.min_column + 14)}{(ws.max_row)}'] = "Indefinido" #QUANTOS SEGUROS RESTANTES DE HELICOPTEROS
    ws[f'{get_column_letter(ws.min_column + 15)}{(ws.max_row)}'] = "❌" #POSSUI ALGUMA CONSTRUÇÃO
    ws[f'{get_column_letter(ws.min_column + 16)}{(ws.max_row)}'] = "0" #QUANTOS DIAS RESTANTES DA CONSTRUÇÃO
    ws[f'{get_column_letter(ws.min_column + 17)}{(ws.max_row)}'] = "❌" #POSSUI FILA PRIORITARIA
    ws[f'{get_column_letter(ws.min_column + 18)}{(ws.max_row)}'] = "0" #QUANTOS DIAS RESTANTES 
    wb.save(file)
    return
    
@bot.event
async def on_ready():
    print("<==========================>")
    print("STATE OF WAR BOT - ONLINE")
    print("<==========================>")
    guild = bot.get_guild(947237264596041728)
    member_count = guild.member_count
    activity = discord.Activity(name="Sobreviventes ➝ ({})".format(member_count), type=1)
    await bot.change_presence(status=discord.Status.online, activity=activity)
    return

@bot.event
async def on_raw_reaction_add(payload):
    ourMessageID = 961364004721291274
    if ourMessageID == payload.message_id:
        member = payload.member
        guild = member.guild
        emoji = payload.emoji.name
        if emoji == '☢️':
            role = discord.utils.get(guild.roles, name="Chernarus")
        elif emoji == '❄️':
            role = discord.utils.get(guild.roles, name="Namalsk")
        await member.add_roles(role)

@bot.event
async def on_raw_reaction_remove(payload):
    ourMessageID = 961364004721291274

    if ourMessageID == payload.message_id:
        guild = await(bot.fetch_guild(payload.guild_id))
        emoji = payload.emoji.name

        if emoji == '☢️':
            role = discord.utils.get(guild.roles, name="Chernarus")
        elif emoji == '❄️':
            role = discord.utils.get(guild.roles, name="Namalsk")
        
        member = await(guild.fetch_member(payload.user_id))
        await member.remove_roles(role)

@clan_criar.error
async def clan_criarerror(ctx, error):
    if isinstance(error, commands.MissingRequiredArgument):
        embedVar = discord.Embed(title="⚠️ Erro na formatação do comando", colour=discord.Colour.random())
        embedVar.add_field(name="Atenção!", value="・Para criar um clan é necessário seguir estritamente o comando `!clan_criar [TAG]`.\n・A cor definida para o clan é aleatória.")
        embedVar.set_footer(text="➝ Se precisar de ajuda, entre em contato com o DEV.")
        await ctx.reply(embed=embedVar)
    return

@doacoesanuncios.error
async def da(ctx, error):
    if isinstance(error, discord.errors.Forbidden):
        embedVar = discord.Embed(title="⚠️ Erro ao enviar as mensagens!", colour = discord.Colour.random())
        embedVar.add_field(name="Atenção:", value=f"<@{ctx.author.id}>, as configurações do seu Discord não permitem que eu te envie as inforamções no seu privado.")
        embedVar.set_footer(text="➝ Se precisar de ajuda, entre em contato com o DEV.")
        message = await ctx.reply(embed=embedVar)
        time.sleep(20)
        await message.delete()
    return

@perfil.error
async def perfil(ctx, error):
    if isinstance(error, discord.errors.Forbidden):
        embedVar = discord.Embed(title="⚠️ As suas configuração não nos permitem te enviar mensagem no privado!", colour = discord.Colour.random())
        embedVar.add_field(name="Atenção:", value="O usuário que você estava tentando convidar não permite receber mensagens privadas de Bots no Discord. Entretanto, não consegui convida-lo!")
        embedVar.set_footer(text="➝ Se precisar de ajuda, entre em contato com o DEV.")
        await ctx.reply(embed = embedVar)
    return

bot.add_cog(Music(bot))
bot.load_extension('cogs.Clan')
bot.run('OTEyMzEwNjYwNjY5NTIxOTIx.YZuFgw.pmM_bRVw0VXQ6OLPrqUu336wpcM')
