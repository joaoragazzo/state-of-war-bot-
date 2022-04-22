import discord 
from discord import channel, colour, reaction 
from discord.ext import commands, tasks 
from discord.ext.commands.core import check, command 
from discord import File 
import discord_components 
from discord_components import DiscordComponents, Select, SelectOption, Button, ButtonStyle, ComponentsBot 
import openpyxl 
from openpyxl import load_workbook 
from openpyxl.utils import get_column_letter 
import time
import datetime

intents = discord.Intents.all()
bot = commands.Bot(command_prefix=['adm '], description="Bot", intents=intents)
DiscordComponents(bot)
client = discord.Client(intents=intents)
update_state = "OFF"

n = 0
@tasks.loop(seconds=10)
async def update_days():
    print("UPDATE DO BANCO DE DADOS INICIADO - COMEÇADO!")

    global update_state

    file = "banco de dados.xlsx"
    wb = openpyxl.load_workbook(filename=file)
    ws = wb.worksheets[0]
    wsr = wb.active

    global n

    if n != 1: #IF para não permitir que desconte 1 dia dos VIPs só por ligar o bot
        n += 1
        return 
    
    update_state = "ON"

    channel = bot.get_channel(960022037794025522)

    async def update(coluna):
        if coluna == 11:
            beneficio = "Seguro de Veiculos"
        if coluna == 14:
            beneficio = "Seguro de Helicopteros"
        if coluna == 17:
            beneficio = "Base VIP"
        if coluna == 19:
            beneficio = "Fila prioritária"

        status = "normal"

        for row in wsr.iter_cols(min_col=coluna, max_col=coluna, min_row=1):
                for cell in row:
                    if cell.value != '0' and cell.value != None:
                        idd = ws.cell(row=cell.row, column=1).value
                        idd = idd.replace("`","")
                        antigo = int(cell.value)
                        novo = antigo - 1
                        ws[f'{get_column_letter(cell.column)}{(cell.row)}'] = str(novo) #TEMPO RESTANTE DA VILA PRIORITARIA
                        if str(novo) == '0':
                            ws[f'{get_column_letter(cell.column -1)}{(cell.row)}'] = '❌'
                            status = "acabou"
                        wb.save(file)
                        if str(novo) == '7':
                            status = "quase"

        if status == "quase":
            user = await bot.fetch_user(idd)
            embedVar = discord.Embed(title="⚠️ Atenção!", description=f"O benefício de `{beneficio}` de {user.name}, portador do Discord ID {idd} está a 7 dias do seu fim", colour=discord.Colour.random())
            embedVar.set_footer(text="Se precisar de ajuda, entre em contato com o DEV.")
            await channel.send(embed = embedVar)

            embedVar = discord.Embed(title="⚠️ Atenção!", description=f"Faltam 7 dias para seu benefício de `{beneficio}` acabar! Você pode renovar criando um ticket na aba de Financeiro!", colour=discord.Colour.random())
            embedVar.set_footer(text="Se precisar de ajuda, entre em contato com o DEV.")
            await user.send(embed = embedVar)
        
        if status == "acabou":
            user = await bot.fetch_user(idd)
            embedVar = discord.Embed(title="⚠️ Atenção!", description=f"O benefício de `{beneficio}` de {user.name}, portador do Discord ID {idd} **acabou**!", colour=discord.Colour.random())
            embedVar.set_footer(text="Se precisar de ajuda, entre em contato com o DEV.")
            await channel.send(embed = embedVar)

            embedVar = discord.Embed(title="⚠️ Atenção!", description=f"O seu beneficio de `{beneficio}` acabou! Você pode renovar criando um ticket na aba de Financeiro!", colour=discord.Colour.random())
            embedVar.set_footer(text="Se precisar de ajuda, entre em contato com o DEV.")
            await user.send(embed = embedVar)
        


    await update(11)
    await update(14)
    await update(17)
    await update(19)

    update_state = "OFF"
    return

@bot.event
async def on_ready():
    print("<==========================>")
    print("STATE OF WAR BOT (ADMINISTRATIVO) - ONLINE")
    print("<==========================>")
    update_days.start()


bot.run()